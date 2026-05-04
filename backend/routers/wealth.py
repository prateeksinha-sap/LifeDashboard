"""
routers/wealth.py
─────────────────────────────────────────────────────────────
GET  /api/wealth            — aggregate net worth + breakdown
POST /api/wealth/manual     — upsert a manual asset (EPF/PPF/NPS/BANK/GOLD_GRAMS)
POST /api/wealth/stocks     — add / update a stock holding
POST /api/wealth/refresh-nav — refresh MF NAVs from mfapi.in
─────────────────────────────────────────────────────────────
"""

import os
import io
import csv
import json
import sys
import math
import re
import xml.etree.ElementTree as ET
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from types import SimpleNamespace
from collections import defaultdict
import time
import tempfile
from datetime import datetime, date, timedelta, timezone
from typing import Optional
from pathlib import Path
from urllib.parse import quote_plus

from sqlalchemy import func

import httpx
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database.db import get_db
from database.models import (
    AssetSnapshot,
    CategoryRule,
    EquitySecurityClassification,
    FundPortfolioStock,
    HistoricalWealth,
    Liability,
    MFHolding,
    ManualAsset,
    MonthClose,
    StockHolding,
    Transaction,
)
from services.mf_nav import search_scheme_code, refresh_nav_for_holding
from services.stock_price import get_stock_price, get_gold_price_inr_per_gram, get_multiple_prices
from services.transaction_categorizer import categorize_transaction_rule, extract_merchant
from services.finance_engine import build_finance_profile
from services.cashflow_rules import HIDDEN_CASHFLOW_CATEGORIES, is_hidden_cashflow_category
from services.equity_sync import refresh_amfi_security_master, sync_equity_lookthrough, sync_mfdata_fund_portfolios
from services.index_lookthrough import get_index_constituents, identify_index_for_security
from services.ai_service import (
    ANTHROPIC_MODEL,
    OLLAMA_MODEL,
    OPENAI_MODEL,
    generate_anthropic_text,
    generate_openai_text,
    generate_text,
    _extract_json,
)

# ── Simple in-memory price cache (5-minute TTL) ───────────────────────
_price_cache: dict[str, float] = {}
_price_cache_ts: float = 0.0
_PRICE_TTL = 300  # seconds

def _safe(v) -> float:
    """Return 0.0 for None, NaN, or Inf — keeps JSON serialization clean."""
    if v is None:
        return 0.0
    try:
        f = float(v)
        return 0.0 if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return 0.0

router = APIRouter(prefix="/api/wealth", tags=["wealth"])

# ── Colour palettes ───────────────────────────────────────────────────
SLICE_COLORS = {
    "Mutual Funds": "#bf5af2",
    "Stocks":       "#0a84ff",
    "EPF":          "#5ac8fa",
    "PPF":          "#30d158",
    "NPS":          "#ff375f",
    "Cash / Bank":  "#ffd60a",
    "Gold":         "#ff9f0a",
    "Fixed Deposits": "#64d2ff",
    "Real Estate":    "#ac8e68",
}

ASSET_TYPE_COLORS = {
    "Equity": "#0a84ff",
    "Debt":   "#5ac8fa",
    "Gold":   "#ff9f0a",
    "Cash":   "#ffd60a",
    "Real Estate": "#ac8e68",
}

INVESTMENT_EXPENSE_CATEGORIES = {"Investments & Savings"}
MUTUAL_FUND_INVESTMENT_KEYWORDS = (
    "sip",
    "mutual fund",
    "mf utility",
    "mfutility",
    "mf util",
    "mf/",
    "cams",
    "kfin",
    "kfintech",
    "amfi",
    "amc",
    "indian clearing corp",
    "iccl",
    "groww",
    "coin",
)

# Keywords that mark a MF scheme as Debt (everything else → Equity)
_DEBT_KEYWORDS = [
    "ultra short", "short duration", "short term fund", "liquid",
    "money market", "overnight", "low duration", "floater",
    "floating rate", "debt", "bond", "gilt", "credit risk",
    "arbitrage", "savings fund", "conservative",
]

def _classify_mf(scheme_name: str) -> str:
    name = scheme_name.lower()
    return "Debt" if any(kw in name for kw in _DEBT_KEYWORDS) else "Equity"


EQUITY_BUCKET_COLORS = {
    "Large Cap": "#0a84ff",
    "Mid Cap": "#5ac8fa",
    "Small Cap": "#bf5af2",
}


EQUITY_CATEGORIES = ("Large Cap", "Mid Cap", "Small Cap")


SECTOR_COLOR_PALETTE = (
    "#30d158",
    "#0a84ff",
    "#bf5af2",
    "#ff9f0a",
    "#64d2ff",
    "#ffd60a",
    "#ff453a",
    "#ac8e68",
    "#5e5ce6",
    "#8e8e93",
    "#00c7be",
    "#ff2d55",
)


_SECTOR_NORMALISATION = {
    "AUTO": "Consumer Cyclical",
    "AUTOMOBILE": "Consumer Cyclical",
    "AUTOMOBILES": "Consumer Cyclical",
    "AUTO ANCILLARIES": "Consumer Cyclical",
    "BANK": "Financial Services",
    "BANKS": "Financial Services",
    "CAPITAL GOODS": "Industrials",
    "CEMENT": "Basic Materials",
    "CHEMICALS": "Basic Materials",
    "COMMODITIES": "Basic Materials",
    "CONSTRUCTION": "Industrials",
    "CONSUMER DURABLES": "Consumer Cyclical",
    "CONSUMER GOODS": "Consumer Defensive",
    "E-RETAIL/ E-COMMERCE": "Consumer Cyclical",
    "ENERGY": "Energy",
    "ETF": "ETF / Index",
    "ETF / INDEX": "ETF / Index",
    "FERTILISERS": "Basic Materials",
    "FERTILIZERS": "Basic Materials",
    "FINANCE": "Financial Services",
    "FINANCIAL SERVICES": "Financial Services",
    "FMCG": "Consumer Defensive",
    "HEALTHCARE": "Healthcare",
    "INDEX DERIVATIVE": "ETF / Index",
    "INFORMATION TECHNOLOGY": "Technology",
    "IT": "Technology",
    "MEDIA": "Communication Services",
    "METAL": "Basic Materials",
    "METALS": "Basic Materials",
    "METALS & MINING": "Basic Materials",
    "MISCELLANEOUS": "Other",
    "OIL": "Energy",
    "OIL & GAS": "Energy",
    "PHARMA": "Healthcare",
    "PHARMACEUTICALS": "Healthcare",
    "POWER": "Utilities",
    "PRIVATE SECTOR BANK": "Financial Services",
    "PUBLIC SECTOR BANK": "Financial Services",
    "REFINERIES & MARKETING": "Energy",
    "REAL ESTATE": "Real Estate",
    "REALTY": "Real Estate",
    "SERVICES": "Industrials",
    "SOFTWARE": "Technology",
    "SOFTWARE SERVICES": "Technology",
    "TELECOM": "Communication Services",
    "TEXTILES": "Consumer Cyclical",
    "TOURISM & HOSPITALITY": "Consumer Cyclical",
    "UTILITIES": "Utilities",
}


def _norm_text(value: str | None) -> str:
    text = str(value or "").lower()
    return "".join(ch for ch in text if ch.isalnum())


def _norm_symbol(value: str | None) -> str:
    text = str(value or "").upper().replace(".NS", "").replace(".BO", "").strip()
    if "-" in text:
        base, suffix = text.split("-", 1)
        if base and suffix in {"T", "XT", "BE", "BZ", "EQ", "SM", "ST"}:
            text = base
    return text


def _security_name_key(value: str | None) -> str:
    text = str(value or "").lower()
    text = re.sub(r"\([^)]*\)|\[[^\]]*\]", " ", text)
    text = text.replace("&", " and ")
    replacements = {
        "corporation": "corp",
        "corpn": "corp",
        "company": "co",
        "laboratories": "labs",
    }
    for src, dest in replacements.items():
        text = text.replace(src, dest)
    drop_words = {
        "ltd", "limited", "ordinary", "shares", "share", "class", "the",
        "india", "indian", "co", "corp", "of", "and", "partly", "paid",
        "rs", "registered", "shs",
    }
    words = [word for word in re.findall(r"[a-z0-9]+", text) if word not in drop_words]
    return "".join(words)


def _normalise_sector(value: str | None) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"na", "n/a", "none", "unknown", "null"}:
        return "Unknown"

    compact = re.sub(r"\s+", " ", text).strip()
    upper = compact.upper()
    if upper in _SECTOR_NORMALISATION:
        return _SECTOR_NORMALISATION[upper]

    for token, sector in _SECTOR_NORMALISATION.items():
        if token in upper and token not in {"IT", "OIL"}:
            return sector

    if upper.startswith("BANK") or " BANK" in upper:
        return "Financial Services"
    if "FINANC" in upper or "BROKING" in upper or "INSURANCE" in upper:
        return "Financial Services"
    if "TECH" in upper or "SOFTWARE" in upper:
        return "Technology"
    if "PHARMA" in upper or "HEALTH" in upper:
        return "Healthcare"
    if "AUTO" in upper:
        return "Consumer Cyclical"
    if "REFINER" in upper or "OIL" in upper or "GAS" in upper or "POWER" in upper:
        return "Energy"
    if "RETAIL" in upper or "HOTEL" in upper or "TOURISM" in upper or "AIRLINE" in upper:
        return "Consumer Cyclical"
    if "WHEELER" in upper or "VEHICLE" in upper or "PASSENGER CAR" in upper or "JEWELL" in upper:
        return "Consumer Cyclical"
    if "FOOD" in upper or "BEVERAGE" in upper or "PERSONAL CARE" in upper or "BREWER" in upper or "DISTILL" in upper or "TEA" in upper or "COFFEE" in upper:
        return "Consumer Defensive"
    if "AEROSPACE" in upper or "DEFENSE" in upper or "DEFENCE" in upper or "ELECTRICAL" in upper or "COMPRESSOR" in upper or "PUMP" in upper or "ENGINE" in upper or "EXPLOSIVE" in upper:
        return "Industrials"
    if "HOSPITAL" in upper:
        return "Healthcare"
    if "HOLDING" in upper or "ASSET MANAGEMENT" in upper or "INVESTMENT COMPANY" in upper:
        return "Financial Services"
    if "RESIDENTIAL" in upper or "COMMERCIAL PROJECT" in upper:
        return "Real Estate"
    if "METAL" in upper or "MINING" in upper or "CHEMICAL" in upper or "CEMENT" in upper or "STEEL" in upper or "ALUMIN" in upper or "COAL" in upper or "ZINC" in upper or "MINERAL" in upper or "PAINT" in upper:
        return "Basic Materials"

    return compact.title() if compact.isupper() else compact


def _known_sector_from_label(value: str | None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    return _SECTOR_NORMALISATION.get(re.sub(r"\s+", " ", text).strip().upper())


def _sector_color(index: int) -> str:
    return SECTOR_COLOR_PALETTE[index % len(SECTOR_COLOR_PALETTE)]


def _synthetic_security(symbol: str | None, name: str | None):
    text = f"{symbol or ''} {name or ''}".lower()
    index_name = identify_index_for_security(symbol, name)
    if index_name:
        category = "Large Cap"
        if index_name == "NIFTY MIDCAP 150":
            category = "Mid Cap"
        elif index_name == "NIFTY SMALLCAP 250":
            category = "Small Cap"
        return SimpleNamespace(category=category, sector="Index ETF", security_name=name or symbol, index_name=index_name)

    global_large_caps = {
        "alphabet": "Communication Services",
        "meta platforms": "Communication Services",
        "amazon.com": "Consumer Cyclical",
        "microsoft": "Technology",
        "cognizant": "Technology",
        "epam systems": "Technology",
        "makemytrip": "Consumer Cyclical",
        "nvidia": "Technology",
        "apple inc": "Technology",
    }
    for token, sector in global_large_caps.items():
        if token in text:
            return SimpleNamespace(category="Large Cap", sector=sector, security_name=name or symbol)
    if "bank nifty" in text or "nifty march" in text or "future on bank index" in text:
        return SimpleNamespace(category="Large Cap", sector="Index derivative", security_name=name or symbol)
    if "niftybees" in text or "nifty 50" in text:
        return SimpleNamespace(category="Large Cap", sector="ETF / Index", security_name=name or symbol)
    if "juniorbees" in text or "nifty next 50" in text:
        return SimpleNamespace(category="Large Cap", sector="ETF / Index", security_name=name or symbol)
    if "midcap" in text or "mid cap" in text:
        return SimpleNamespace(category="Mid Cap", sector="ETF / Index", security_name=name or symbol)
    if "smallcap" in text or "small cap" in text:
        return SimpleNamespace(category="Small Cap", sector="ETF / Index", security_name=name or symbol)
    return None


def _normalise_equity_category(value: str | None) -> str | None:
    text = str(value or "").strip().lower().replace("-", " ").replace("_", " ")
    if not text:
        return None
    if "large" in text and "mid" not in text:
        return "Large Cap"
    if "mid" in text:
        return "Mid Cap"
    if "small" in text or "micro" in text:
        return "Small Cap"
    if text in {"large cap", "largecap"}:
        return "Large Cap"
    if text in {"mid cap", "midcap"}:
        return "Mid Cap"
    if text in {"small cap", "smallcap"}:
        return "Small Cap"
    return None


def _normalise_fund_name(value: str | None) -> str:
    text = str(value or "").lower()
    drops = [
        "direct plan", "regular plan", "direct", "regular", "growth option",
        "growth plan", "growth", "idcw", "payout", "reinvestment", "option",
        "plan", "-", "(", ")", "&",
    ]
    for token in drops:
        text = text.replace(token, " ")
    words = [word for word in text.split() if word not in {"the", "fund", "scheme"}]
    return "".join(words)


def _read_tabular_upload(content: bytes, filename: str) -> list[dict[str, object]]:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(500, f"Excel import requires openpyxl: {exc}")
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        rows: list[list[object]] = []
        for ws in wb.worksheets:
            rows.extend([list(row) for row in ws.iter_rows(values_only=True)])
        return _rows_to_dicts(rows)

    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return _rows_to_dicts([row for row in reader])


def _rows_to_dicts(rows: list[list[object]]) -> list[dict[str, object]]:
    header_idx = None
    for idx, row in enumerate(rows[:40]):
        normalised = {_norm_header(cell) for cell in row if cell not in (None, "")}
        if normalised and (
            {"stockname", "companyname", "nameofcompany", "securityname"} & normalised
            or {"scheme", "schemename", "fund", "fundname"} & normalised
        ):
            header_idx = idx
            break
    if header_idx is None:
        return []
    headers = [str(cell or "").strip() for cell in rows[header_idx]]
    result: list[dict[str, object]] = []
    for row in rows[header_idx + 1:]:
        if not any(str(cell or "").strip() for cell in row):
            continue
        values = list(row) + [None] * max(0, len(headers) - len(row))
        result.append({headers[i]: values[i] for i in range(len(headers)) if headers[i]})
    return result


def _row_value(row: dict[str, object], *names: str) -> object | None:
    normalised = {_norm_header(key): value for key, value in row.items()}
    for name in names:
        if name in normalised:
            return normalised[name]
    return None


def _parse_date_optional(value: object | None) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%b-%Y", "%B-%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue
    return None


# ── Pydantic schemas ──────────────────────────────────────────────────
class ManualAssetIn(BaseModel):
    asset_type: str   # EPF | PPF | NPS | BANK | GOLD_GRAMS
    value: float
    notes: Optional[str] = None


class StockHoldingIn(BaseModel):
    symbol:       str
    company_name: Optional[str] = None
    quantity:     int
    avg_price:    float


class CategoryCorrectionIn(BaseModel):
    transaction_id: int
    category: str
    scope: str = "merchant"  # transaction | merchant


def _norm_header(s) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("_", "").replace("%", "").replace(".", "")


def _parse_num(s) -> float:
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return _safe(s)
    text = (
        str(s)
        .replace(",", "")
        .replace("INR", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .strip()
    )
    if text in ("", "-", "nan", "None"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _saved_category_for_transaction(db: Session, row: Transaction) -> tuple[str | None, float]:
    merchant = extract_merchant(row.description).lower()
    description = str(row.description or "").lower()
    rules = (
        db.query(CategoryRule)
        .filter(CategoryRule.transaction_type == row.transaction_type)
        .order_by(CategoryRule.updated_at.desc())
        .all()
    )
    for rule in rules:
        pattern = str(rule.pattern or "").lower().strip()
        if pattern and (pattern in merchant or pattern in description):
            return str(rule.category), 0.98
    return None, 0.0


def _category_for_transaction(db: Session, row: Transaction) -> tuple[str, float, bool]:
    """Return the effective category, confidence, and whether it was derived."""
    rule_category, rule_confidence = _saved_category_for_transaction(db, row)
    saved_category = (row.category or "").strip()
    if rule_category:
        return rule_category, rule_confidence, False
    derived_category, derived_confidence = categorize_transaction_rule(row.description, row.transaction_type)
    if (
        saved_category
        and saved_category.lower() not in {"misc", "miscellaneous", "uncategorized", "transfers out"}
        and not (saved_category == "Food & Delivery" and derived_category == "Household Help")
    ):
        return saved_category, 1.0, False
    return derived_category, derived_confidence, True


def _is_mutual_fund_investment(row: Transaction, category: str) -> bool:
    """True when an investment debit looks like a MF/SIP transfer."""
    if category not in INVESTMENT_EXPENSE_CATEGORIES:
        return False
    text = f"{row.description or ''} {extract_merchant(row.description)}".lower()
    return any(keyword in text for keyword in MUTUAL_FUND_INVESTMENT_KEYWORDS)


def _column_lookup(headers: list[str], *candidates: str) -> int | None:
    normalised = {_norm_header(h): i for i, h in enumerate(headers)}
    for candidate in candidates:
        idx = normalised.get(candidate)
        if idx is not None:
            return idx
    return None


def _parse_stock_rows(rows: list[list]) -> tuple[list[dict], int, list[str]]:
    if not rows:
        return [], 0, []

    header_idx = None
    for idx, row in enumerate(rows):
        normalised = {_norm_header(cell) for cell in row}
        if "symbol" in normalised and (
            "quantityavailable" in normalised
            or "qty" in normalised
            or "quantity" in normalised
            or "shares" in normalised
        ):
            header_idx = idx
            break

    if header_idx is None:
        preview = [str(c) for row in rows[:8] for c in row if c not in (None, "")]
        return [], 0, preview[:16]

    headers = [str(h or "").strip() for h in rows[header_idx]]
    symbol_idx = _column_lookup(headers, "symbol", "instrument", "tradingsymbol", "stock")
    qty_idx = _column_lookup(headers, "quantityavailable", "qty", "quantity", "shares")
    avg_idx = _column_lookup(headers, "averageprice", "avgcost", "avgprice", "purchaseprice", "buyprice")
    ltp_idx = _column_lookup(headers, "previousclosingprice", "ltp", "lastprice", "currentprice", "closingprice", "price")
    name_idx = _column_lookup(headers, "companyname", "name", "company")
    sector_idx = _column_lookup(headers, "sector")

    if symbol_idx is None or qty_idx is None or avg_idx is None:
        return [], 0, headers

    parsed: list[dict] = []
    skipped = 0
    for raw in rows[header_idx + 1:]:
        row = list(raw) + [None] * max(0, len(headers) - len(raw))
        symbol = str(row[symbol_idx] or "").strip().upper()
        if not symbol or symbol in ("SYMBOL", "INSTRUMENT", "TOTAL"):
            skipped += 1
            continue

        qty = int(_parse_num(row[qty_idx]))
        if qty <= 0:
            skipped += 1
            continue

        avg_price = _parse_num(row[avg_idx])
        ltp = _parse_num(row[ltp_idx]) if ltp_idx is not None else 0.0
        name = str(row[name_idx] or "").strip() if name_idx is not None else None
        sector = str(row[sector_idx] or "").strip() if sector_idx is not None else None

        parsed.append({
            "symbol": symbol,
            "quantity": qty,
            "avg_price": avg_price,
            "current_price": ltp,
            "company_name": name or sector or None,
        })

    return parsed, skipped, headers


def _parse_stocks_csv(content: bytes) -> tuple[list[dict], int, list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    return _parse_stock_rows(rows)


def _parse_stocks_xlsx(content: bytes) -> tuple[list[dict], int, list[str], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(500, f"Excel import requires openpyxl: {exc}")

    # Zerodha's workbook declares dimension=A1 even though data exists below it.
    # read_only mode trusts that bad dimension, so load normally and let openpyxl
    # calculate the actual B4:M42-style range.
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    preferred = ["Equity", "Combined", *wb.sheetnames]
    seen: set[str] = set()
    best_error: list[str] = []
    for sheet_name in preferred:
        if sheet_name in seen or sheet_name not in wb.sheetnames:
            continue
        seen.add(sheet_name)
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        parsed, skipped, headers = _parse_stock_rows(rows)
        if parsed:
            return parsed, skipped, headers, sheet_name
        best_error = headers
    return [], 0, best_error, preferred[0] if preferred else "workbook"


def _upsert_stock_holdings(db: Session, holdings: list[dict]) -> tuple[int, int]:
    imported = updated = 0
    for item in holdings:
        symbol = item["symbol"]
        holding = db.query(StockHolding).filter_by(symbol=symbol).first()
        if holding:
            holding.quantity      = item["quantity"]
            holding.avg_price     = item["avg_price"]
            holding.current_price = item["current_price"] if item["current_price"] > 0 else holding.current_price
            holding.company_name  = item["company_name"] or holding.company_name
            holding.updated_at    = datetime.utcnow()
            updated += 1
        else:
            db.add(StockHolding(
                symbol=symbol,
                company_name=item["company_name"],
                quantity=item["quantity"],
                avg_price=item["avg_price"],
                current_price=item["current_price"] if item["current_price"] > 0 else None,
            ))
        imported += 1

    if imported > 0:
        month_year = date.today().strftime("%Y-%m")
        close = db.query(MonthClose).filter_by(month_year=month_year).first()
        if not close:
            close = MonthClose(month_year=month_year)
            db.add(close)
        close.investments_refreshed = True

    db.commit()
    return imported, updated


# ── Helpers ───────────────────────────────────────────────────────────
def _gold_value_inr(grams: float) -> float:
    env_price = _safe(os.getenv("GOLD_PRICE_INR_PER_GRAM"))
    if env_price > 0:
        return round(grams * env_price, 2)
    if os.getenv("LIVE_GOLD_PRICE", "").strip().lower() not in {"1", "true", "yes"}:
        return round(grams * 9000, 2)
    price = get_gold_price_inr_per_gram()
    if price:
        return round(grams * price, 2)
    # Fallback: use approximate ₹9,000/gram if yfinance unavailable
    return round(grams * 9000, 2)


# ── Endpoints ─────────────────────────────────────────────────────────
@router.get("")
def get_wealth(db: Session = Depends(get_db)):
    """
    Aggregate net worth from:
      1. MF holdings (CAS import)
      2. Stock holdings (Zerodha import / saved prices)
      3. Manual assets (bank cash, FD, PF, PPF, NPS, gold, real estate)
    """
    # 1. Mutual Funds
    mf_holdings = db.query(MFHolding).all()
    mf_value = round(sum(h.value for h in mf_holdings), 2)

    # 2. Stocks - use saved Zerodha/imported prices for fast dashboard loads.
    # Live network refresh is available through the explicit refresh endpoint.
    stocks = db.query(StockHolding).all()
    stock_value = round(sum((_safe(s.current_price) or _safe(s.avg_price)) * s.quantity for s in stocks), 2)

    # 3. Manual assets
    manual = {a.asset_type: a.value for a in db.query(ManualAsset).all()}
    epf_value  = manual.get("EPF",        0)
    ppf_value  = manual.get("PPF",        0)
    nps_value  = manual.get("NPS",        0)
    bank_value = manual.get("BANK",       0)
    fd_value   = manual.get("FD",         0)
    real_estate_value = manual.get("REAL_ESTATE", 0)
    gold_grams = manual.get("GOLD_GRAMS", 0)
    gold_value = _gold_value_inr(gold_grams) if gold_grams > 0 else 0

    gross_assets = mf_value + stock_value + epf_value + ppf_value + nps_value + bank_value + fd_value + real_estate_value + gold_value
    liabilities_total = round(sum(_safe(row.outstanding_amount) for row in db.query(Liability).all()), 2)
    total = round(gross_assets - liabilities_total, 2)

    # Build slices (skip zero-value slices)
    raw_slices = [
        ("Mutual Funds", mf_value),
        ("Stocks",       stock_value),
        ("EPF",          epf_value),
        ("PPF",          ppf_value),
        ("NPS",          nps_value),
        ("Cash / Bank",  bank_value),
        ("Fixed Deposits", fd_value),
        ("Real Estate", real_estate_value),
        ("Gold",         gold_value),
    ]

    slices = []
    for label, value in raw_slices:
        if value > 0 and gross_assets > 0:
            slices.append({
                "label":      label,
                "value":      value,
                "percentage": round((value / gross_assets) * 100, 1),
                "color":      SLICE_COLORS.get(label, "#888"),
            })

    # ── Asset-type breakdown (Equity / Debt / Gold / Cash) ────────────
    equity_mf = sum(h.value for h in mf_holdings if _classify_mf(h.scheme_name) == "Equity")
    debt_mf   = sum(h.value for h in mf_holdings if _classify_mf(h.scheme_name) == "Debt")

    asset_buckets = {
        "Equity": round(equity_mf + stock_value,              2),
        "Debt":   round(debt_mf + epf_value + ppf_value + nps_value + fd_value, 2),
        "Gold":   round(gold_value,                            2),
        "Cash":   round(bank_value,                            2),
        "Real Estate": round(real_estate_value,                 2),
    }
    asset_type_slices = [
        {
            "label":      label,
            "value":      value,
            "percentage": round((value / gross_assets) * 100, 1) if gross_assets else 0,
            "color":      ASSET_TYPE_COLORS[label],
        }
        for label, value in asset_buckets.items()
        if value > 0
    ]

    return {
        "total_net_worth":   total,
        "gross_assets":      round(gross_assets, 2),
        "liabilities":       liabilities_total,
        "slices":            slices,
        "asset_type_slices": asset_type_slices,
        "mf_count":          len(mf_holdings),
        "stock_count":       len(stocks),
        "last_updated":      datetime.utcnow().isoformat(),
    }


@router.get("/equity-allocation")
def get_equity_allocation(db: Session = Depends(get_db)):
    """Return Large/Mid/Small look-through equity exposure."""
    mf_holdings = db.query(MFHolding).all()
    stock_holdings = db.query(StockHolding).all()
    security_rows = db.query(EquitySecurityClassification).all()
    portfolio_rows = db.query(FundPortfolioStock).all()

    by_symbol = {_norm_symbol(row.symbol): row for row in security_rows if row.symbol}
    by_isin = {str(row.isin or "").upper().strip(): row for row in security_rows if row.isin}
    by_name = {_norm_text(row.security_name): row for row in security_rows if row.security_name}
    by_name_key = {_security_name_key(row.security_name): row for row in security_rows if row.security_name}
    fuzzy_cache: dict[str, object | None] = {}
    portfolios_by_fund: dict[str, list[FundPortfolioStock]] = defaultdict(list)
    for row in portfolio_rows:
        portfolios_by_fund[_normalise_fund_name(row.scheme_name)].append(row)

    def lookup_security(symbol: str | None, isin: str | None, name: str | None):
        synthetic = _synthetic_security(symbol, name)
        if synthetic:
            return synthetic
        symbol_key = _norm_symbol(symbol)
        if symbol_key and symbol_key in by_symbol:
            return by_symbol[symbol_key]
        if isin and str(isin).upper().strip() in by_isin:
            return by_isin[str(isin).upper().strip()]
        if name:
            key = _norm_text(name)
            if key in by_name:
                return by_name[key]
            name_key = _security_name_key(name)
            if name_key in by_name_key:
                return by_name_key[name_key]
            if len(name_key) >= 5:
                for candidate_key, row in by_name_key.items():
                    if not candidate_key:
                        continue
                    shorter = min(len(name_key), len(candidate_key))
                    longer = max(len(name_key), len(candidate_key))
                    if shorter / longer >= 0.45 and (name_key in candidate_key or candidate_key in name_key):
                        return row
            if len(name_key) < 8:
                return None
            if name_key in fuzzy_cache:
                return fuzzy_cache[name_key]
            best = None
            best_score = 0.0
            for candidate_key, row in by_name_key.items():
                if not candidate_key or not name_key:
                    continue
                score = SequenceMatcher(None, name_key, candidate_key).ratio()
                if score > best_score:
                    best = row
                    best_score = score
            fuzzy_cache[name_key] = best if best_score >= 0.88 else None
            return fuzzy_cache[name_key]
        return None

    bucket_totals = {label: 0.0 for label in EQUITY_CATEGORIES}
    bucket_rows: dict[str, list[dict]] = {label: [] for label in EQUITY_CATEGORIES}
    exposure_rows: list[dict] = []
    unmapped_rows: list[dict] = []
    missing_funds: list[dict] = []

    total_equity = 0.0

    def register_equity_row(row: dict, category: str | None) -> None:
        if category in EQUITY_CATEGORIES:
            bucket_totals[category] += _safe(row.get("value"))
            bucket_rows[category].append(row)
            exposure_rows.append(row)
        else:
            row["reason"] = row.get("reason") or "Stock not mapped to AMFI large/mid/small category"
            unmapped_rows.append(row)

    def expand_index_row(
        *,
        base_row: dict,
        index_name: str,
        exposure_value: float,
        parent_weight_pct: float,
    ) -> bool:
        constituents = get_index_constituents(index_name)
        if not constituents:
            base_row["category"] = None
            base_row["sector"] = "Index ETF - constituents unavailable"
            base_row["reason"] = f"Could not fetch constituents for {index_name}"
            unmapped_rows.append(base_row)
            return False

        for constituent in constituents:
            child_weight = max(_safe(constituent.get("weight_pct")), 0.0)
            if child_weight <= 0:
                continue
            security = lookup_security(
                constituent.get("symbol"),
                constituent.get("isin"),
                constituent.get("stock_name"),
            )
            category = _normalise_equity_category(
                security.category if security else constituent.get("category")
            )
            sector = _normalise_sector(
                security.sector if security and security.sector else constituent.get("sector")
            )
            child_value = round(exposure_value * child_weight / 100, 2)
            child_row = {
                "type": "ETF Look-through",
                "source": f"{base_row['source']} -> {index_name}",
                "stock_name": constituent.get("stock_name") or constituent.get("symbol") or index_name,
                "symbol": constituent.get("symbol"),
                "isin": constituent.get("isin"),
                "category": category,
                "sector": sector,
                "weight_pct": round(child_weight, 4),
                "parent_weight_pct": round(parent_weight_pct, 4),
                "value": child_value,
                "fund_value": base_row.get("fund_value"),
                "lookthrough_source": constituent.get("source"),
                "weight_method": constituent.get("weight_method"),
            }
            register_equity_row(child_row, category)
        return True

    for holding in mf_holdings:
        if _classify_mf(holding.scheme_name) != "Equity":
            continue
        value = round(_safe(holding.value), 2)
        if value <= 0:
            continue
        total_equity += value
        matched = portfolios_by_fund.get(_normalise_fund_name(holding.scheme_name), [])
        if not matched:
            missing_funds.append({
                "type": "Mutual Fund",
                "source": holding.scheme_name,
                "value": value,
                "reason": "Fund portfolio composition not imported",
            })
            continue
        imported_weight = 0.0
        for portfolio in matched:
            weight = max(_safe(portfolio.weight_pct), 0.0)
            if weight <= 0:
                continue
            imported_weight += weight
            security = lookup_security(portfolio.symbol, portfolio.isin, portfolio.stock_name)
            category = _normalise_equity_category(
                security.category if security else portfolio.category
            )
            sector = _normalise_sector((security.sector if security and security.sector else portfolio.sector) or None)
            exposure_value = round(value * weight / 100, 2)
            row = {
                "type": "Mutual Fund",
                "source": holding.scheme_name,
                "stock_name": portfolio.stock_name,
                "symbol": portfolio.symbol,
                "isin": portfolio.isin,
                "category": category,
                "sector": sector,
                "weight_pct": round(weight, 4),
                "value": exposure_value,
                "fund_value": value,
            }
            index_name = getattr(security, "index_name", None) or identify_index_for_security(portfolio.symbol, portfolio.stock_name)
            if index_name:
                expand_index_row(
                    base_row=row,
                    index_name=index_name,
                    exposure_value=exposure_value,
                    parent_weight_pct=weight,
                )
                continue
            register_equity_row(row, category)
        if imported_weight < 95:
            missing_funds.append({
                "type": "Mutual Fund",
                "source": holding.scheme_name,
                "value": round(value * max(0, 100 - imported_weight) / 100, 2),
                "reason": f"Only {round(imported_weight, 1)}% of fund portfolio imported",
            })

    for stock in stock_holdings:
        price = _safe(stock.current_price) or _safe(stock.avg_price)
        value = round(price * _safe(stock.quantity), 2)
        if value <= 0:
            continue
        total_equity += value
        direct_sector = _known_sector_from_label(stock.company_name)
        security = lookup_security(stock.symbol, None, stock.symbol)
        category = _normalise_equity_category(security.category if security else None)
        sector = _normalise_sector(security.sector if security and security.sector else direct_sector)
        stock_name = (
            (security.security_name if security and getattr(security, "security_name", None) else None)
            or (stock.company_name if stock.company_name and not direct_sector else None)
            or stock.symbol
        )
        row = {
            "type": "Direct Stock",
            "source": stock.symbol,
            "stock_name": stock_name,
            "symbol": stock.symbol,
            "isin": None,
            "category": category,
            "sector": sector,
            "weight_pct": 100.0,
            "quantity": stock.quantity,
            "avg_price": _safe(stock.avg_price),
            "current_price": price,
            "value": value,
            "updated_at": stock.updated_at.isoformat() if stock.updated_at else None,
        }
        index_name = getattr(security, "index_name", None) or identify_index_for_security(stock.symbol, stock.company_name)
        if index_name:
            expand_index_row(
                base_row=row,
                index_name=index_name,
                exposure_value=value,
                parent_weight_pct=100.0,
            )
            continue
        if category not in EQUITY_CATEGORIES:
            row["reason"] = "Direct stock not mapped to AMFI large/mid/small category"
        register_equity_row(row, category)

    total_equity = round(total_equity, 2)
    mapped_value = round(sum(bucket_totals.values()), 2)
    unmapped_value = round(
        sum(row["value"] for row in unmapped_rows) + sum(row["value"] for row in missing_funds),
        2,
    )
    buckets = []
    for label in EQUITY_CATEGORIES:
        value = round(bucket_totals[label], 2)
        items = sorted(bucket_rows[label], key=lambda row: row["value"], reverse=True)
        buckets.append({
            "label": label,
            "value": value,
            "percentage": round((value / total_equity) * 100, 1) if total_equity else 0,
            "color": EQUITY_BUCKET_COLORS[label],
            "count": len(items),
            "rows": items,
        })

    sector_totals: dict[str, float] = defaultdict(float)
    sector_counts: dict[str, int] = defaultdict(int)
    sector_unmapped_value = 0.0
    for row in [*exposure_rows, *unmapped_rows]:
        value = _safe(row.get("value"))
        if value <= 0:
            continue
        sector = _normalise_sector(row.get("sector"))
        if sector in {"Unknown", "ETF / Index"} or "unavailable" in sector.lower():
            sector_unmapped_value += value
            continue
        sector_totals[sector] += value
        sector_counts[sector] += 1

    missing_sector_value = round(sum(row["value"] for row in missing_funds), 2)
    if missing_sector_value > 0:
        sector_unmapped_value += missing_sector_value

    sorted_sectors = sorted(sector_totals.items(), key=lambda item: item[1], reverse=True)
    sector_mapped_value = round(sum(sector_totals.values()), 2)
    sector_allocation = [
        {
            "label": label,
            "value": round(value, 2),
            "percentage": round((value / total_equity) * 100, 1) if total_equity else 0,
            "count": sector_counts[label],
            "color": _sector_color(idx),
        }
        for idx, (label, value) in enumerate(sorted_sectors)
    ]

    return {
        "total_equity": total_equity,
        "mapped_value": mapped_value,
        "unmapped_value": unmapped_value,
        "coverage_pct": round((mapped_value / total_equity) * 100, 1) if total_equity else 0,
        "bucket_count": 3,
        "sector_count": len(sector_allocation),
        "sector_mapped_value": sector_mapped_value,
        "sector_unmapped_value": round(sector_unmapped_value, 2),
        "sector_coverage_pct": round((sector_mapped_value / total_equity) * 100, 1) if total_equity else 0,
        "row_count": len(exposure_rows),
        "holding_count": len(mf_holdings) + len(stock_holdings),
        "security_master_count": len(security_rows),
        "fund_portfolio_row_count": len(portfolio_rows),
        "classification_method": (
            "Look-through exposure: fund value multiplied by imported stock weights, "
            "then mapped to AMFI/SEBI Large/Mid/Small categories. Unmapped exposure is kept out of the three buckets."
        ),
        "buckets": buckets,
        "sector_allocation": sector_allocation,
        "rows": sorted(exposure_rows, key=lambda row: row["value"], reverse=True),
        "unmapped": {
            "value": unmapped_value,
            "missing_fund_composition": sorted(missing_funds, key=lambda row: row["value"], reverse=True),
            "unclassified_rows": sorted(unmapped_rows, key=lambda row: row["value"], reverse=True),
        },
        "last_updated": datetime.utcnow().isoformat(),
    }


SECTOR_GUIDANCE_SOURCES = {
    "Reuters",
    "Bloomberg",
    "Business Standard",
    "Mint",
    "The Economic Times",
    "Moneycontrol",
    "CNBC TV18",
    "Morningstar",
    "Value Research",
    "Motilal Oswal",
    "ICICI Direct",
    "HDFC Securities",
    "Kotak Securities",
    "Axis Securities",
    "Nuvama",
    "JM Financial",
    "RBI",
    "NSE",
}

_SECTOR_SOURCE_CACHE: dict[int, tuple[float, list[dict]]] = {}
_SECTOR_GUIDANCE_CACHE: dict[tuple, tuple[float, dict]] = {}


def _cache_minutes(name: str, default: int) -> int:
    try:
        return max(1, int(os.getenv(name, str(default))))
    except ValueError:
        return default


def _copy_dict_list(items: list[dict]) -> list[dict]:
    return [dict(item) for item in items]


def _sector_guidance_signature(equity: dict) -> tuple:
    return (
        round(float(equity.get("total_equity") or 0), 0),
        round(float(equity.get("sector_coverage_pct") or 0), 1),
        tuple(
            (row.get("label"), round(float(row.get("value") or 0), 0), round(float(row.get("percentage") or 0), 1))
            for row in (equity.get("sector_allocation") or [])[:12]
        ),
    )


SECTOR_KEYWORDS = {
    "Financial Services": ("bank", "banks", "nbfc", "financial", "insurance", "credit", "lending"),
    "Technology": ("it", "technology", "software", "ai", "digital", "semiconductor"),
    "Healthcare": ("healthcare", "pharma", "pharmaceutical", "hospital", "diagnostic"),
    "Consumer Cyclical": ("auto", "retail", "consumer discretionary", "travel", "hotel", "ecommerce", "e-commerce"),
    "Consumer Defensive": ("fmcg", "consumer staples", "food", "rural demand"),
    "Industrials": ("industrial", "capital goods", "infrastructure", "manufacturing", "defence", "defense"),
    "Basic Materials": ("metal", "mining", "cement", "chemical", "materials"),
    "Energy": ("energy", "oil", "gas", "power", "renewable", "utilities"),
    "Communication Services": ("telecom", "media", "communication"),
    "Real Estate": ("real estate", "property", "reit", "housing"),
}


def _sector_mentions(text: str) -> list[str]:
    lower = text.lower()
    mentions: list[str] = []
    for sector, keywords in SECTOR_KEYWORDS.items():
        matched = False
        for keyword in keywords:
            if len(keyword) <= 3:
                matched = re.search(rf"\b{re.escape(keyword)}\b", lower) is not None
            else:
                matched = keyword in lower
            if matched:
                break
        if matched:
            mentions.append(sector)
    return mentions


def _parse_rss_date(value: str | None) -> str | None:
    if not value:
        return None
    try:
        parsed = parsedate_to_datetime(value)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


async def _fetch_sector_guidance_sources(days: int) -> list[dict]:
    ttl = _cache_minutes("SECTOR_NEWS_CACHE_MINUTES", 360) * 60
    cached = _SECTOR_SOURCE_CACHE.get(days)
    now_ts = time.time()
    if cached and now_ts - cached[0] < ttl:
        return _copy_dict_list(cached[1])

    queries = [
        "India equity market sector outlook overweight underweight banks IT pharma consumption capital goods energy 2026",
        "India stock market sector outlook banks pharma IT capital goods consumption 2026",
        "India brokerage sector strategy overweight underweight financials IT healthcare capital goods 2026",
        "India mutual fund manager sector outlook banks IT pharma consumption energy 2026",
    ]
    try:
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
            responses = []
            for query in queries:
                url = f"https://news.google.com/rss/search?q={quote_plus(query + f' when:{days}d')}&hl=en-IN&gl=IN&ceid=IN:en"
                response = await client.get(url, headers={"User-Agent": "LifeDashboard/1.0"})
                response.raise_for_status()
                responses.append(response.text)
    except Exception:
        return []

    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    sources: list[dict] = []
    seen: set[str] = set()
    for response_text in responses:
        root = ET.fromstring(response_text)
        for item in root.findall(".//item")[:80]:
            title = item.findtext("title") or ""
            summary = item.findtext("description") or ""
            source = item.findtext("source") or "Google News"
            published_iso = _parse_rss_date(item.findtext("pubDate"))
            if published_iso:
                try:
                    if datetime.fromisoformat(published_iso) < cutoff:
                        continue
                except ValueError:
                    pass
            if source not in SECTOR_GUIDANCE_SOURCES and not any(name.lower() in title.lower() for name in SECTOR_GUIDANCE_SOURCES):
                continue
            sectors = _sector_mentions(f"{title} {summary}")
            if not sectors:
                continue
            key = f"{source}:{title}"
            if key in seen:
                continue
            seen.add(key)
            sources.append({
                "title": title,
                "url": item.findtext("link") or "",
                "source": source,
                "published_iso": published_iso,
                "sectors": sectors,
            })
            if len(sources) >= 14:
                break
        if len(sources) >= 14:
            break
    _SECTOR_SOURCE_CACHE[days] = (now_ts, _copy_dict_list(sources))
    return sources


def _deterministic_sector_guidance(equity: dict, sources: list[dict]) -> dict:
    sectors = equity.get("sector_allocation") or []
    sector_by_name = {row["label"]: row for row in sectors}
    mention_counts: dict[str, int] = defaultdict(int)
    for source in sources:
        for sector in source.get("sectors", []):
            mention_counts[sector] += 1

    suggestions: list[dict] = []
    for sector in sectors[:8]:
        pct_value = float(sector.get("percentage") or 0)
        mentions = mention_counts.get(sector["label"], 0)
        if pct_value >= 22:
            stance = "reduce"
            action = "Do not add fresh money here until your next SIP review; redirect new contributions to underweight diversified sectors."
            why = f"{sector['label']} is already {pct_value:.1f}% of equity exposure."
        elif mentions >= 2 and pct_value < 8:
            stance = "add"
            action = "Research diversified funds or existing holdings that add exposure gradually, not a single-stock bet."
            why = f"{sector['label']} is only {pct_value:.1f}% of equity exposure and appears in recent sector-outlook sources."
        else:
            stance = "hold"
            action = "Keep current exposure; review after the next monthly snapshot or if sources flag a clear risk."
            why = f"{sector['label']} exposure is {pct_value:.1f}%."
        suggestions.append({
            "sector": sector["label"],
            "stance": stance,
            "why": why,
            "action": action,
            "confidence": "medium" if sources else "low",
            "dashboard_evidence": [
                f"Exposure {sector['percentage']:.1f}%",
                f"Value {sector['value']:.0f}",
                f"Stock rows {sector['count']}",
            ],
            "source_evidence": [
                src["title"] for src in sources
                if sector["label"] in src.get("sectors", [])
            ][:3],
        })

    for sector, count in sorted(mention_counts.items(), key=lambda item: item[1], reverse=True):
        if sector in sector_by_name:
            continue
        suggestions.append({
            "sector": sector,
            "stance": "research",
            "why": f"{sector} appears in {count} recent reliable-source item(s), but current app exposure is not visible.",
            "action": "Check whether this sector is hidden inside unmapped funds before adding anything.",
            "confidence": "low",
            "dashboard_evidence": ["No mapped exposure in current sector tile"],
            "source_evidence": [src["title"] for src in sources if sector in src.get("sectors", [])][:3],
        })
        if len(suggestions) >= 8:
            break

    return {
        "headline": "Sector guidance is based on your mapped exposure and recent reliable-source mentions.",
        "suggestions": suggestions[:8],
        "limitations": [
            "This is not a buy/sell recommendation.",
            "Source headlines are screened for sector relevance; read the linked articles before acting.",
            "Unmapped fund or ETF exposure can change the conclusion.",
        ],
    }


def _sector_guidance_prompt(equity: dict, sources: list[dict]) -> str:
    context = {
        "as_of": datetime.now(timezone.utc).isoformat(),
        "sector_allocation": equity.get("sector_allocation", [])[:12],
        "sector_coverage_pct": equity.get("sector_coverage_pct"),
        "total_equity": equity.get("total_equity"),
        "sources": sources[:14],
    }
    return f"""You are an India-focused portfolio analyst.

Use ONLY the dashboard exposure and source list below. Do not invent sources, returns, or stock tips.

Return ONLY JSON with this shape:
{{
  "headline": "one practical sentence",
  "suggestions": [
    {{
      "sector": "sector name",
      "stance": "add" | "hold" | "reduce" | "research",
      "why": "short reason combining portfolio exposure and source evidence",
      "action": "what the user should review or do next",
      "confidence": "low" | "medium" | "high",
      "dashboard_evidence": ["facts from allocation"],
      "source_evidence": ["source title 1", "source title 2"]
    }}
  ],
  "limitations": ["short limitation"]
}}

Rules:
- Do not recommend buying or selling individual stocks.
- "Add" means consider future diversified allocation, not immediate trade.
- "Reduce" means reduce new contributions or rebalance at review time, not panic sell.
- Prefer 5-7 high-signal suggestions.
- Mention when sector coverage is incomplete.

FACTS:
{json.dumps(context, ensure_ascii=True, indent=2)}
"""


async def _ai_sector_guidance(equity: dict, sources: list[dict]) -> tuple[dict, str, str | None]:
    prompt = _sector_guidance_prompt(equity, sources)
    provider = os.getenv("SECTOR_GUIDANCE_PROVIDER", os.getenv("COACH_PROVIDER", os.getenv("ASSISTANT_PROVIDER", "auto"))).strip().lower()
    has_openai = bool(os.getenv("OPENAI_API_KEY"))
    has_anthropic = bool(os.getenv("ANTHROPIC_API_KEY"))
    errors: list[Exception] = []

    if provider in {"openai", "premium", "auto"} and has_openai:
        try:
            text = await generate_openai_text(prompt, max_tokens=1400, model=OPENAI_MODEL, reasoning_effort="medium")
            return _extract_json(text), "openai", OPENAI_MODEL
        except Exception as exc:
            errors.append(exc)
    if provider in {"anthropic", "claude", "auto"} and has_anthropic:
        try:
            text = await generate_anthropic_text(prompt, max_tokens=1400, model=ANTHROPIC_MODEL)
            return _extract_json(text), "anthropic", ANTHROPIC_MODEL
        except Exception as exc:
            errors.append(exc)
    if provider in {"ollama", "local", "auto"}:
        try:
            text = await generate_text(prompt, max_tokens=1400, temperature=0.1, model=OLLAMA_MODEL)
            return _extract_json(text), "ollama", OLLAMA_MODEL
        except Exception as exc:
            errors.append(exc)
    if has_openai:
        try:
            text = await generate_openai_text(prompt, max_tokens=1400, model=OPENAI_MODEL, reasoning_effort="medium")
            return _extract_json(text), "openai", OPENAI_MODEL
        except Exception as exc:
            errors.append(exc)
    if has_anthropic:
        try:
            text = await generate_anthropic_text(prompt, max_tokens=1400, model=ANTHROPIC_MODEL)
            return _extract_json(text), "anthropic", ANTHROPIC_MODEL
        except Exception as exc:
            errors.append(exc)
    if errors:
        raise errors[-1]
    raise RuntimeError("No LLM configured for sector guidance")


@router.get("/equity-allocation/sector-guidance")
async def get_sector_guidance(
    days: int = Query(default=45, ge=7, le=120),
    use_ai: bool = Query(default=True),
    force: bool = Query(default=False),
    db: Session = Depends(get_db),
):
    """Source-grounded sector guidance based on current look-through exposure."""
    equity = get_equity_allocation(db)
    provider_setting = os.getenv("SECTOR_GUIDANCE_PROVIDER", os.getenv("COACH_PROVIDER", os.getenv("ASSISTANT_PROVIDER", "auto"))).strip().lower()
    cache_key = (days, bool(use_ai), provider_setting, _sector_guidance_signature(equity))
    ttl = _cache_minutes("SECTOR_GUIDANCE_CACHE_MINUTES", 360) * 60
    now_ts = time.time()
    cached = _SECTOR_GUIDANCE_CACHE.get(cache_key)
    if cached and not force and now_ts - cached[0] < ttl:
        result = dict(cached[1])
        result["cached"] = True
        return result

    sources = await _fetch_sector_guidance_sources(days)
    generated_at = datetime.now(timezone.utc).isoformat()

    if use_ai and sources:
        try:
            guidance, provider, model = await _ai_sector_guidance(equity, sources)
            result = {
                "status": "ok",
                "generated_at": generated_at,
                "provider": provider,
                "model": model,
                "fallback": False,
                "cached": False,
                "source_count": len(sources),
                "sources": sources,
                **guidance,
            }
            _SECTOR_GUIDANCE_CACHE[cache_key] = (now_ts, dict(result))
            return result
        except Exception:
            pass

    fallback = _deterministic_sector_guidance(equity, sources)
    result = {
        "status": "ok",
        "generated_at": generated_at,
        "provider": "deterministic",
        "model": None,
        "fallback": True,
        "cached": False,
        "source_count": len(sources),
        "sources": sources,
        **fallback,
    }
    _SECTOR_GUIDANCE_CACHE[cache_key] = (now_ts, dict(result))
    return result


@router.post("/equity-allocation/import-security-master")
async def import_equity_security_master(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import AMFI/SEBI large-mid-small security classification."""
    content = await file.read()
    rows = _read_tabular_upload(content, file.filename or "")
    if not rows:
        raise HTTPException(400, "Could not find a usable table in this file.")

    imported = updated = skipped = 0
    current_section: str | None = None
    for raw in rows:
        first_values = " ".join(str(v or "") for v in list(raw.values())[:3]).lower()
        section_category = _normalise_equity_category(first_values)
        if section_category and len([v for v in raw.values() if str(v or "").strip()]) <= 2:
            current_section = section_category
            skipped += 1
            continue

        name = str(_row_value(raw, "stockname", "companyname", "nameofcompany", "securityname", "name") or "").strip()
        symbol = _norm_symbol(_row_value(raw, "symbol", "nsesymbol", "nse", "ticker", "stocksymbol"))
        isin = str(_row_value(raw, "isin", "isinno", "isinumber") or "").upper().strip() or None
        sector = str(_row_value(raw, "sector", "industry", "basicindustry") or "").strip() or None
        category = _normalise_equity_category(_row_value(raw, "category", "marketcapcategory", "capcategory", "classification")) or current_section

        if not category:
            rank = _parse_num(_row_value(raw, "rank", "srno", "serialno", "sno"))
            if rank > 0:
                if rank <= 100:
                    category = "Large Cap"
                elif rank <= 250:
                    category = "Mid Cap"
                else:
                    category = "Small Cap"

        if not name or category not in EQUITY_CATEGORIES:
            skipped += 1
            continue

        existing = None
        if symbol:
            existing = db.query(EquitySecurityClassification).filter_by(symbol=symbol).first()
        if not existing and isin:
            existing = db.query(EquitySecurityClassification).filter_by(isin=isin).first()
        if not existing:
            existing = db.query(EquitySecurityClassification).filter_by(security_name=name, symbol=symbol or None, isin=isin).first()

        if existing:
            existing.security_name = name
            existing.symbol = symbol or existing.symbol
            existing.isin = isin or existing.isin
            existing.category = category
            existing.sector = sector or existing.sector
            existing.source = file.filename
            existing.updated_at = datetime.utcnow()
            updated += 1
        else:
            db.add(EquitySecurityClassification(
                security_name=name,
                symbol=symbol or None,
                isin=isin,
                category=category,
                sector=sector,
                source=file.filename,
            ))
            imported += 1
    db.commit()
    return {"status": "ok", "imported": imported, "updated": updated, "skipped": skipped}


@router.post("/equity-allocation/import-fund-portfolio")
async def import_fund_portfolio(
    file: UploadFile = File(...),
    scheme_name: str = Form(default=""),
    as_of_date: str = Form(default=""),
    db: Session = Depends(get_db),
):
    """Import mutual-fund underlying stock weights from factsheet/portfolio exports."""
    content = await file.read()
    rows = _read_tabular_upload(content, file.filename or "")
    if not rows:
        raise HTTPException(400, "Could not find a usable table in this file.")

    imported = updated = skipped = 0
    parsed_date = _parse_date_optional(as_of_date)
    for raw in rows:
        scheme = str(_row_value(raw, "scheme", "schemename", "fund", "fundname", "portfolio") or scheme_name or "").strip()
        stock_name = str(_row_value(raw, "stockname", "companyname", "nameofcompany", "securityname", "instrument", "holding", "name") or "").strip()
        symbol = _norm_symbol(_row_value(raw, "symbol", "nsesymbol", "nse", "ticker", "stocksymbol"))
        isin = str(_row_value(raw, "isin", "isinno", "isinumber") or "").upper().strip() or None
        sector = str(_row_value(raw, "sector", "industry", "basicindustry") or "").strip() or None
        amc = str(_row_value(raw, "amc", "fundhouse") or "").strip() or None
        category = _normalise_equity_category(_row_value(raw, "category", "marketcapcategory", "capcategory", "classification"))
        row_date = _parse_date_optional(_row_value(raw, "date", "asofdate", "portfolio date")) or parsed_date
        weight = _parse_num(_row_value(
            raw,
            "weight", "weightpct", "weightage", "holdingpct", "percentage",
            "tonav", "tonetassets", "pctonav", "netasset", "netassets",
            "aumpct", "portfolioholding", "assetallocation",
        ))

        if weight > 1_000:
            weight = 0
        if 0 < weight <= 1:
            # Some exports use 0.062 for 6.2%.
            weight *= 100

        if not scheme or not stock_name or weight <= 0:
            skipped += 1
            continue

        existing = (
            db.query(FundPortfolioStock)
            .filter_by(scheme_name=scheme, stock_name=stock_name, symbol=symbol or None, isin=isin)
            .first()
        )
        if existing:
            existing.amc = amc or existing.amc
            existing.category = category or existing.category
            existing.sector = sector or existing.sector
            existing.weight_pct = weight
            existing.as_of_date = row_date or existing.as_of_date
            existing.source = file.filename
            existing.updated_at = datetime.utcnow()
            updated += 1
        else:
            db.add(FundPortfolioStock(
                scheme_name=scheme,
                amc=amc,
                stock_name=stock_name,
                symbol=symbol or None,
                isin=isin,
                category=category,
                sector=sector,
                weight_pct=weight,
                as_of_date=row_date,
                source=file.filename,
            ))
            imported += 1
    db.commit()
    return {"status": "ok", "imported": imported, "updated": updated, "skipped": skipped}


@router.post("/equity-allocation/refresh-security-master")
async def refresh_equity_security_master(db: Session = Depends(get_db)):
    """Refresh AMFI official large/mid/small stock classification automatically."""
    try:
        return await refresh_amfi_security_master(db)
    except Exception as exc:
        db.rollback()
        raise HTTPException(502, f"AMFI classification refresh failed: {exc}")


@router.post("/equity-allocation/sync-fund-portfolios")
async def sync_equity_fund_portfolios(limit: int | None = None, db: Session = Depends(get_db)):
    """Resolve CAS mutual funds and import latest stock-level portfolios from mfdata.in."""
    try:
        return await sync_mfdata_fund_portfolios(db, limit=limit)
    except Exception as exc:
        db.rollback()
        raise HTTPException(502, f"Mutual fund portfolio sync failed: {exc}")


@router.post("/equity-allocation/sync")
async def sync_full_equity_lookthrough(limit: int | None = None, db: Session = Depends(get_db)):
    """Refresh AMFI classifications and MF look-through holdings in one call."""
    try:
        return await sync_equity_lookthrough(db, limit=limit)
    except Exception as exc:
        db.rollback()
        raise HTTPException(502, f"Equity look-through sync failed: {exc}")


@router.get("/manual")
def get_manual_assets(db: Session = Depends(get_db)):
    """Return all manual asset balances as a flat dict."""
    assets = db.query(ManualAsset).all()
    result = {a.asset_type: a.value for a in assets}
    for key in ("EPF", "PPF", "NPS", "BANK", "GOLD_GRAMS", "FD", "REAL_ESTATE"):
        result.setdefault(key, 0)
    return result


@router.post("/manual")
def upsert_manual_asset(body: ManualAssetIn, db: Session = Depends(get_db)):
    """Create or update a manual asset balance."""
    asset = db.query(ManualAsset).filter_by(asset_type=body.asset_type).first()
    if asset:
        asset.value      = body.value
        asset.notes      = body.notes
        asset.updated_at = datetime.utcnow()
    else:
        db.add(ManualAsset(
            asset_type=body.asset_type,
            value=body.value,
            notes=body.notes,
        ))
    db.commit()
    return {"status": "ok", "asset_type": body.asset_type, "value": body.value}


@router.get("/stocks")
def list_stocks(db: Session = Depends(get_db)):
    """Return all stock holdings with live prices."""
    holdings = db.query(StockHolding).all()
    result = []
    for s in holdings:
        price    = _safe(s.current_price) or _safe(s.avg_price)
        avg      = _safe(s.avg_price)
        result.append({
            "id":            s.id,
            "symbol":        s.symbol,
            "company_name":  s.company_name,
            "quantity":      s.quantity,
            "avg_price":     avg,
            "current_price": price,
            "value":         round(price * s.quantity, 2),
            "gain_loss":     round((price - avg) * s.quantity, 2),
            "gain_pct":      round(((price - avg) / avg) * 100, 2) if avg else 0,
        })
    return result


@router.post("/stocks")
def upsert_stock(body: StockHoldingIn, db: Session = Depends(get_db)):
    """Add or update a Zerodha stock holding."""
    holding = db.query(StockHolding).filter_by(symbol=body.symbol.upper()).first()
    if holding:
        holding.quantity     = body.quantity
        holding.avg_price    = body.avg_price
        holding.company_name = body.company_name
        holding.updated_at   = datetime.utcnow()
    else:
        db.add(StockHolding(
            symbol=body.symbol.upper(),
            company_name=body.company_name,
            quantity=body.quantity,
            avg_price=body.avg_price,
        ))
    db.commit()
    return {"status": "ok", "symbol": body.symbol.upper()}


@router.delete("/stocks/{symbol}")
def delete_stock(symbol: str, db: Session = Depends(get_db)):
    holding = db.query(StockHolding).filter_by(symbol=symbol.upper()).first()
    if not holding:
        raise HTTPException(404, "Stock not found")
    db.delete(holding)
    db.commit()
    return {"status": "ok"}


@router.post("/stocks/import")
async def import_stocks_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import Zerodha holdings from CSV or the official Console XLSX export."""
    content = await file.read()
    if not content:
        raise HTTPException(400, "Empty file")

    filename = (file.filename or "").lower()
    source_sheet = None
    if filename.endswith((".xlsx", ".xlsm")):
        holdings, skipped, headers, source_sheet = _parse_stocks_xlsx(content)
        import_type = "xlsx"
    elif filename.endswith((".csv", ".txt")):
        holdings, skipped, headers = _parse_stocks_csv(content)
        import_type = "csv"
    else:
        raise HTTPException(400, "Please upload a Zerodha holdings .xlsx or .csv file.")

    if not holdings:
        raise HTTPException(
            400,
            "Could not find stock holdings. Expected columns like Symbol, Quantity Available, "
            f"Average Price, Previous Closing Price. Found: {headers}",
        )

    imported, updated = _upsert_stock_holdings(db, holdings)
    return {
        "status": "ok",
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "format": import_type,
        "sheet": source_sheet,
    }


@router.post("/stocks/import-csv-legacy")
async def import_stocks_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import Zerodha holdings CSV.
    Zerodha Console export columns:
      Instrument, Qty, Avg cost, LTP, Cur val, P&L, Net chg%, Day chg%
    Also handles alternate headers like: Symbol, Quantity, Average Price, ...
    """
    content = await file.read()
    text = content.decode("utf-8-sig", errors="replace")  # strip BOM if present
    reader = csv.DictReader(io.StringIO(text))

    # Normalise header names
    def norm(s: str) -> str:
        return s.strip().lower().replace(" ", "").replace("_", "").replace("%", "")

    rows = list(reader)
    if not rows:
        raise HTTPException(400, "Empty CSV file")

    # Map flexible column names
    sample = {norm(k): k for k in rows[0].keys()}

    def col(*candidates):
        for c in candidates:
            if c in sample:
                return sample[c]
        return None

    symbol_col  = col("instrument", "symbol", "tradingsymbol", "stock")
    qty_col     = col("qty", "quantity", "shares")
    avgcost_col = col("avgcost", "avgprice", "averageprice", "purchaseprice", "buyprice")
    ltp_col     = col("ltp", "lastprice", "currentprice", "closingprice", "price")
    name_col    = col("companyname", "name", "company")

    if not symbol_col or not qty_col or not avgcost_col:
        raise HTTPException(400,
            f"Could not find required columns. Found: {list(rows[0].keys())}. "
            "Expected: Instrument/Symbol, Qty/Quantity, Avg cost/Average Price")

    def parse_num(s: str) -> float:
        return float(str(s).replace(",", "").replace("₹", "").strip() or 0)

    imported, skipped = 0, 0
    for row in rows:
        symbol = str(row.get(symbol_col, "")).strip().upper()
        if not symbol or symbol in ("", "INSTRUMENT", "SYMBOL"):
            skipped += 1
            continue

        qty = int(parse_num(row.get(qty_col, 0)))
        if qty <= 0:
            skipped += 1
            continue

        avg_price = parse_num(row.get(avgcost_col, 0))
        ltp       = parse_num(row.get(ltp_col, 0)) if ltp_col else 0
        name      = str(row.get(name_col, "")).strip() if name_col else None

        holding = db.query(StockHolding).filter_by(symbol=symbol).first()
        if holding:
            holding.quantity      = qty
            holding.avg_price     = avg_price
            holding.current_price = ltp if ltp > 0 else holding.current_price
            holding.company_name  = name or holding.company_name
            holding.updated_at    = datetime.utcnow()
        else:
            db.add(StockHolding(
                symbol=symbol,
                company_name=name,
                quantity=qty,
                avg_price=avg_price,
                current_price=ltp if ltp > 0 else None,
            ))
        imported += 1

    db.commit()
    return {"status": "ok", "imported": imported, "skipped": skipped}


@router.post("/stocks/refresh-prices")
async def refresh_stock_prices(db: Session = Depends(get_db)):
    """Refresh saved stock prices from Yahoo Finance on explicit user request."""
    holdings = db.query(StockHolding).all()
    if not holdings:
        return {"status": "ok", "updated": 0, "total": 0, "skipped": 0}

    prices = get_multiple_prices([h.symbol for h in holdings])
    updated = 0
    for holding in holdings:
        price = _safe(prices.get(holding.symbol))
        if price > 0:
            holding.current_price = price
            holding.updated_at = datetime.utcnow()
            updated += 1

    db.commit()
    return {
        "status": "ok",
        "updated": updated,
        "total": len(holdings),
        "skipped": len(holdings) - updated,
    }


@router.post("/refresh-nav")
async def refresh_mf_nav(db: Session = Depends(get_db)):
    """
    Refresh live NAV for all MF holdings from mfapi.in.
    Automatically resolves scheme_code if not yet stored.
    Returns count of successfully updated holdings.
    """
    holdings = db.query(MFHolding).all()
    updated = 0

    for h in holdings:
        # Resolve scheme_code once if not stored
        if not h.scheme_code:
            code = await search_scheme_code(h.scheme_name)
            if code:
                h.scheme_code = code

        if h.scheme_code:
            result = await refresh_nav_for_holding(h.scheme_code, h.units)
            if result:
                h.nav        = result["nav"]
                h.value      = result["value"]
                h.nav_date   = datetime.utcnow().date()
                h.updated_at = datetime.utcnow()
                updated += 1

    db.commit()
    return {"status": "ok", "updated": updated, "total": len(holdings)}


@router.post("/mutual-funds/import-cas")
async def import_mutual_funds_cas(
    file: UploadFile = File(...),
    password: str = Form(default=""),
    db: Session = Depends(get_db),
):
    """Upload CAMS/KFintech CAS PDF and import active mutual fund holdings."""
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Please upload the CAS PDF file.")

    cas_password = password.strip() or os.getenv("CAS_PASSWORD", "").strip()
    if not cas_password:
        raise HTTPException(
            status_code=400,
            detail="CAS password is required. Enter it in the upload panel or set CAS_PASSWORD in backend .env.",
        )

    scripts_dir = str(Path(__file__).resolve().parents[1] / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)

    try:
        from import_cas import import_cas
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"CAS import module unavailable: {e}")

    contents = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        try:
            imported = import_cas(tmp_path, cas_password, db)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not import CAS PDF: {e}")
    finally:
        os.unlink(tmp_path)

    if imported > 0:
        month_year = date.today().strftime("%Y-%m")
        close = db.query(MonthClose).filter_by(month_year=month_year).first()
        if not close:
            close = MonthClose(month_year=month_year)
            db.add(close)
        close.investments_refreshed = True
        close.updated_at = datetime.utcnow()
        db.commit()

    return {"status": "ok", "imported": imported, "filename": file.filename}


# ════════════════════════════════════════════════════════════════════
#  Financial Analytics — Trends, Savings Rate, Forecast, Snapshot
# ════════════════════════════════════════════════════════════════════

def _last_n_months(n: int, anchor: date | None = None) -> list[str]:
    """Return n month strings ending at anchor's month, oldest first."""
    today = anchor or date.today()
    months = []
    for i in range(n - 1, -1, -1):
        m = (today.month - i - 1) % 12 + 1
        y = today.year - ((i - today.month + 1 + 12) // 12 if (today.month - i - 1) < 0 else 0)
        months.append(f"{y}-{m:02d}")
    return months


def _month_range(start: date, end: date) -> list[str]:
    """Return YYYY-MM strings from start's month to end's month, inclusive."""
    months = []
    year, month = start.year, start.month
    while (year, month) <= (end.year, end.month):
        months.append(f"{year}-{month:02d}")
        month += 1
        if month == 13:
            month = 1
            year += 1
    return months


def _month_income_expenses(
    db: Session,
    month_year: str,
    *,
    exclude_investments: bool = False,
) -> tuple[float, float]:
    """Sum Credit and Debit transactions for a given YYYY-MM.

    Expenses always hide cash withdrawals. When exclude_investments is true,
    debit transactions categorized as investment/savings transfers are also
    left out so the chart reflects consumption spend.
    """
    income = db.query(func.sum(Transaction.amount)).filter(
        Transaction.transaction_type == "Credit",
        func.strftime("%Y-%m", Transaction.date) == month_year,
    ).scalar() or 0.0

    debit_rows = (
        db.query(Transaction)
        .filter(
            Transaction.transaction_type == "Debit",
            func.strftime("%Y-%m", Transaction.date) == month_year,
        )
        .all()
    )
    expenses = 0.0
    for row in debit_rows:
        category = _category_for_transaction(db, row)[0]
        if is_hidden_cashflow_category(category):
            continue
        if exclude_investments and category in INVESTMENT_EXPENSE_CATEGORIES:
            continue
        expenses += _safe(row.amount)

    return round(float(income), 2), round(float(expenses), 2)


def _month_investment_split(db: Session, month_year: str) -> tuple[float, float, float]:
    """Return total, MF/SIP, and other investment debits for a YYYY-MM month."""
    debit_rows = (
        db.query(Transaction)
        .filter(
            Transaction.transaction_type == "Debit",
            func.strftime("%Y-%m", Transaction.date) == month_year,
        )
        .all()
    )
    mf_investment = 0.0
    other_investment = 0.0
    for row in debit_rows:
        category, _, _ = _category_for_transaction(db, row)
        if category not in INVESTMENT_EXPENSE_CATEGORIES:
            continue
        amount = _safe(row.amount)
        if _is_mutual_fund_investment(row, category):
            mf_investment += amount
        else:
            other_investment += amount

    total = mf_investment + other_investment
    return round(total, 2), round(mf_investment, 2), round(other_investment, 2)


def _median_positive(values: list[float]) -> float:
    positive = sorted(float(value) for value in values if value > 0)
    if not positive:
        return 0.0
    mid = len(positive) // 2
    if len(positive) % 2:
        return positive[mid]
    return (positive[mid - 1] + positive[mid]) / 2


def _latest_cashflow_months(db: Session, limit: int = 6) -> list[str]:
    month_expr = func.strftime("%Y-%m", Transaction.date)
    rows = (
        db.query(month_expr.label("month"))
        .group_by(month_expr)
        .order_by(month_expr.desc())
        .limit(limit)
        .all()
    )
    return [row[0] for row in reversed(rows) if row[0]]


def _current_net_worth_simple(db: Session) -> float:
    """Fast net-worth estimate from DB values (no live yfinance calls)."""
    manual   = {a.asset_type: a.value for a in db.query(ManualAsset).all()}
    mf_val   = sum(h.value for h in db.query(MFHolding).all())
    stk_val  = sum(
        (_safe(s.current_price) or _safe(s.avg_price)) * s.quantity
        for s in db.query(StockHolding).all()
    )
    gold_g   = manual.get("GOLD_GRAMS", 0)
    gold_val = _gold_value_inr(gold_g) if gold_g > 0 else 0
    return round(
        mf_val + stk_val +
        manual.get("EPF", 0) + manual.get("PPF", 0) +
        manual.get("NPS", 0) + manual.get("BANK", 0) +
        manual.get("FD", 0) + manual.get("REAL_ESTATE", 0) + gold_val,
        2,
    )


def _asset_values_simple(db: Session) -> dict[str, float]:
    manual = {a.asset_type: a.value for a in db.query(ManualAsset).all()}
    mf_val = sum(h.value for h in db.query(MFHolding).all())
    stock_val = sum((_safe(s.current_price) or _safe(s.avg_price)) * s.quantity for s in db.query(StockHolding).all())
    gold_g = manual.get("GOLD_GRAMS", 0)
    gold_val = _gold_value_inr(gold_g) if gold_g > 0 else 0
    return {
        "Cash": round(manual.get("BANK", 0), 2),
        "Mutual Funds": round(mf_val, 2),
        "Stocks": round(stock_val, 2),
        "Gold": round(gold_val, 2),
        "Real Estate": round(manual.get("REAL_ESTATE", 0), 2),
        "Fixed Deposits": round(manual.get("FD", 0), 2),
        "PPF": round(manual.get("PPF", 0), 2),
        "PF": round(manual.get("EPF", 0), 2),
        "NPS": round(manual.get("NPS", 0), 2),
    }


@router.get("/trends")
def get_wealth_trends(range: str = "auto", db: Session = Depends(get_db)):
    """
    Returns 12 months of income, expenses, and net-worth snapshots.
    Net worth comes from historical_wealth snapshots if available;
    falls back to the latest known value for months without a snapshot.
    """
    tx_stats = db.query(
        func.count(Transaction.id),
        func.min(Transaction.date),
        func.max(Transaction.date),
    ).one()
    total_tx = int(tx_stats[0] or 0)
    earliest_tx = tx_stats[1]
    latest_tx = tx_stats[2]
    snapshots = db.query(
        func.count(HistoricalWealth.id),
        func.min(HistoricalWealth.month_year),
        func.max(HistoricalWealth.month_year),
    ).one()
    total_snapshots = int(snapshots[0] or 0)
    latest_nw = _current_net_worth_simple(db)

    rolling_months = _last_n_months(12)
    if range == "all" and earliest_tx and latest_tx:
        months = _month_range(earliest_tx, latest_tx)
        range_label = "All imported cashflow"
    elif range == "imported" and latest_tx:
        months = _last_n_months(12, latest_tx)
        range_label = "Imported period"
    else:
        months = rolling_months
        range_label = "Last 12 months"

    rolling_has_cashflow = any(sum(_month_income_expenses(db, my)) > 0 for my in rolling_months)
    if range == "auto" and not rolling_has_cashflow and earliest_tx and latest_tx:
        months = _last_n_months(12, latest_tx)
        range_label = "Imported period"

    result = []
    today = date.today()
    current_month = today.strftime("%Y-%m")
    salary_baseline = _safe(os.getenv("USER_MONTHLY_SALARY_INR")) or 0.0
    salary_signal_threshold = max(salary_baseline * 0.5, 50_000.0)
    provisional_months = []
    for my in months:
        income, expenses = _month_income_expenses(db, my)
        _, expenses_excluding_investments = _month_income_expenses(db, my, exclude_investments=True)
        snapshot = db.query(HistoricalWealth).filter_by(month_year=my).first()
        is_current_month = my == current_month
        has_salary_signal = income >= salary_signal_threshold
        is_provisional = bool(is_current_month and today.day < 25 and not has_salary_signal)
        if is_provisional:
            provisional_months.append(my)
        result.append({
            "month":     my,
            "income":    income,
            "expenses":  expenses,
            "expenses_excluding_investments": expenses_excluding_investments,
            "investment_outflow": round(max(expenses - expenses_excluding_investments, 0), 2),
            "net_worth": round(snapshot.total_net_worth, 2) if snapshot else (
                latest_nw if my == current_month else None
            ),
            "has_data":  income > 0 or expenses > 0,
            "is_current_month": is_current_month,
            "is_provisional": is_provisional,
            "visible_in_trend": not is_provisional or range == "all",
        })

    has_visible_cashflow = any(r["has_data"] for r in result)

    return {
        "months": result,
        "has_transactions": has_visible_cashflow,
        "range_mode": range,
        "range_label": range_label,
        "total_transaction_count": total_tx,
        "earliest_transaction_date": earliest_tx.isoformat() if earliest_tx else None,
        "latest_transaction_date": latest_tx.isoformat() if latest_tx else None,
        "total_snapshot_count": total_snapshots,
        "earliest_snapshot_month": snapshots[1],
        "latest_snapshot_month": snapshots[2],
        "is_showing_imported_period": range_label == "Imported period",
        "provisional_months": provisional_months,
        "provisional_month_note": (
            f"{', '.join(provisional_months)} is hidden from the trend until salary/month-end because the current month is incomplete."
            if provisional_months and range != "all" else None
        ),
        "expense_modes": {
            "all": "Expenses include visible bank debits, including investments and savings transfers. Cash withdrawals are always hidden.",
            "spend_only": "Expenses exclude transactions categorized as Investments & Savings. Cash withdrawals are always hidden.",
        },
    }


@router.get("/transactions/month-breakdown/{month_year}")
def get_month_transaction_breakdown(
    month_year: str,
    direction: str = "Debit",
    exclude_investments: bool = False,
    db: Session = Depends(get_db),
):
    """Category and transaction drill-down for one clicked cashflow bar."""
    try:
        year_s, month_s = month_year.split("-", 1)
        year = int(year_s)
        month = int(month_s)
        if year < 2000 or not 1 <= month <= 12:
            raise ValueError
    except ValueError:
        raise HTTPException(status_code=400, detail="month_year must be YYYY-MM")

    direction_norm = direction.strip().capitalize()
    if direction_norm not in ("Credit", "Debit"):
        raise HTTPException(status_code=400, detail="direction must be Credit or Debit")

    rows = (
        db.query(Transaction)
        .filter(
            Transaction.transaction_type == direction_norm,
            func.strftime("%Y-%m", Transaction.date) == month_year,
        )
        .order_by(Transaction.amount.desc(), Transaction.date.desc())
        .all()
    )

    category_totals: dict[str, dict] = defaultdict(lambda: {"total": 0.0, "count": 0, "merchants": defaultdict(float)})
    transactions = []
    derived_count = 0
    excluded_total = 0.0
    excluded_count = 0
    cash_withdrawals_hidden = False

    for row in rows:
        category, confidence, was_derived = _category_for_transaction(db, row)
        merchant = extract_merchant(row.description)
        amount = round(_safe(row.amount), 2)

        if direction_norm == "Debit" and is_hidden_cashflow_category(category):
            cash_withdrawals_hidden = True
            continue

        if direction_norm == "Debit" and exclude_investments and category in INVESTMENT_EXPENSE_CATEGORIES:
            excluded_total += amount
            excluded_count += 1
            continue

        if was_derived:
            derived_count += 1

        bucket = category_totals[category]
        bucket["total"] += amount
        bucket["count"] += 1
        bucket["merchants"][merchant] += amount

        transactions.append({
            "id": row.id,
            "date": row.date.isoformat(),
            "description": row.description,
            "merchant": merchant,
            "amount": amount,
            "category": category,
            "account_source": row.account_source,
            "confidence": round(confidence, 2),
        })

    total = round(sum(item["amount"] for item in transactions), 2)
    categories = []
    for category, bucket in category_totals.items():
        merchants = sorted(bucket["merchants"].items(), key=lambda item: item[1], reverse=True)
        categories.append({
            "category": category,
            "total": round(bucket["total"], 2),
            "count": bucket["count"],
            "percentage": round((bucket["total"] / total) * 100, 1) if total else 0,
            "top_merchants": [
                {"merchant": merchant, "total": round(value, 2)}
                for merchant, value in merchants[:5]
            ],
        })
    categories.sort(key=lambda item: item["total"], reverse=True)

    return {
        "month": month_year,
        "direction": direction_norm,
        "label": "Income" if direction_norm == "Credit" else (
            "Expenses excl. investments" if exclude_investments else "Expenses"
        ),
        "total": total,
        "transaction_count": len(transactions),
        "raw_transaction_count": len(rows),
        "excluded_category_total": round(excluded_total, 2),
        "excluded_category_count": excluded_count,
        "cash_withdrawals_hidden": cash_withdrawals_hidden,
        "hidden_categories": sorted(HIDDEN_CASHFLOW_CATEGORIES),
        "exclude_investments": bool(direction_norm == "Debit" and exclude_investments),
        "derived_category_count": derived_count,
        "categories": categories,
        "transactions": transactions[:100],
    }


@router.post("/transactions/category-correction")
def save_category_correction(body: CategoryCorrectionIn, db: Session = Depends(get_db)):
    """Remember a user's category correction and apply it to similar transactions."""
    category = body.category.strip()
    if not category:
        raise HTTPException(status_code=400, detail="category is required")

    txn = db.query(Transaction).filter_by(id=body.transaction_id).first()
    if not txn:
        raise HTTPException(status_code=404, detail="transaction not found")

    merchant = extract_merchant(txn.description).strip()
    pattern = merchant if body.scope == "merchant" else str(txn.description or "").strip()
    pattern_norm = pattern.lower()
    if not pattern_norm:
        raise HTTPException(status_code=400, detail="could not derive a reusable pattern")

    rule = (
        db.query(CategoryRule)
        .filter_by(pattern=pattern_norm, transaction_type=txn.transaction_type)
        .first()
    )
    if not rule:
        rule = CategoryRule(
            pattern=pattern_norm,
            category=category,
            transaction_type=txn.transaction_type,
        )
        db.add(rule)
    else:
        rule.category = category
        rule.updated_at = datetime.utcnow()

    matched = (
        db.query(Transaction)
        .filter(Transaction.transaction_type == txn.transaction_type)
        .all()
    )
    updated = 0
    for row in matched:
        row_merchant = extract_merchant(row.description).lower()
        row_description = str(row.description or "").lower()
        if pattern_norm in row_merchant or pattern_norm in row_description:
            row.category = category
            updated += 1

    rule.match_count = updated
    db.commit()
    return {
        "status": "saved",
        "pattern": pattern,
        "category": category,
        "transaction_type": txn.transaction_type,
        "updated": updated,
    }


@router.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, db: Session = Depends(get_db)):
    txn = db.query(Transaction).filter_by(id=transaction_id).first()
    if not txn:
        raise HTTPException(status_code=404, detail="transaction not found")
    db.delete(txn)
    db.commit()
    return {"status": "deleted", "transaction_id": transaction_id}


@router.get("/category-rules")
def get_category_rules(db: Session = Depends(get_db)):
    rows = db.query(CategoryRule).order_by(CategoryRule.updated_at.desc()).all()
    return [
        {
            "id": row.id,
            "pattern": row.pattern,
            "category": row.category,
            "transaction_type": row.transaction_type,
            "match_count": row.match_count,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
        for row in rows
    ]


@router.get("/asset-trends")
def get_asset_trends(db: Session = Depends(get_db)):
    """Monthly asset snapshot trends captured during month close."""
    rows = db.query(AssetSnapshot).order_by(AssetSnapshot.month_year.asc(), AssetSnapshot.asset_type.asc()).all()
    months: dict[str, dict] = {}
    for row in rows:
        months.setdefault(row.month_year, {"month": row.month_year})
        months[row.month_year][row.asset_type] = round(row.value or 0, 2)
    return {
        "months": list(months.values()),
        "asset_types": sorted({row.asset_type for row in rows}),
        "has_snapshots": bool(rows),
    }


@router.get("/savings-rate")
def get_savings_rate(db: Session = Depends(get_db)):
    """
    Current-month savings rate and trailing 6-month average.
    Returns zero rates when no transaction data exists yet.
    """
    today    = date.today()
    curr_my  = f"{today.year}-{today.month:02d}"

    income_now, expenses_now = _month_income_expenses(db, curr_my)
    savings_now = income_now - expenses_now
    rate_now    = (savings_now / income_now * 100) if income_now > 0 else 0.0

    # Trailing 6 months (excluding current)
    past_months = _last_n_months(7)[:-1]   # 6 complete past months
    total_inc_6m = total_exp_6m = 0.0
    for my in past_months:
        inc, exp = _month_income_expenses(db, my)
        total_inc_6m += inc
        total_exp_6m += exp

    avg_savings_6m = (total_inc_6m - total_exp_6m) / 6
    rate_6m        = (
        (total_inc_6m - total_exp_6m) / total_inc_6m * 100
        if total_inc_6m > 0 else 0.0
    )

    return {
        "current_month":              curr_my,
        "income_this_month":          income_now,
        "expenses_this_month":        expenses_now,
        "savings_this_month":         round(savings_now, 2),
        "savings_rate_pct":           round(rate_now, 1),
        "trailing_6m_avg_income":     round(total_inc_6m / 6, 2),
        "trailing_6m_avg_expenses":   round(total_exp_6m / 6, 2),
        "trailing_6m_avg_savings":    round(avg_savings_6m, 2),
        "trailing_6m_savings_rate_pct": round(rate_6m, 1),
        "has_data":                   income_now > 0 or total_inc_6m > 0,
    }


@router.get("/forecast")
def get_wealth_forecast(step_up_pct: float = 10.0, db: Session = Depends(get_db)):
    profile = build_finance_profile(db, step_up_pct=step_up_pct)
    forecast = profile["forecast"]
    return {
        **forecast,
        "assumptions": profile["assumptions"],
        "asset_return_components": profile["assets"]["returns"]["components"],
        "analyzed_months": profile["cashflow"]["analyzed_months"],
    }

    """
    5-year wealth projection.
    Assumes 8% annualised return (compounded monthly) + trailing-6m avg savings.
    Falls back to ₹30,000/month savings if no transaction history exists.
    """
    # Forecast starts from the current live estimate; historical snapshots are
    # for trends, not for today's projection base.
    current_nw = _current_net_worth_simple(db)
    manual = {a.asset_type: a.value for a in db.query(ManualAsset).all()}
    current_cash = round(float(manual.get("BANK", 0) or 0), 2)
    current_invested_assets = round(max(current_nw - current_cash, 0), 2)
    step_up_pct = max(0.0, min(float(step_up_pct or 0), 100.0))

    today = date.today()
    current_month = today.strftime("%Y-%m")
    candidate_months = _latest_cashflow_months(db, 7)
    all_month_rows = []
    for my in candidate_months:
        income, all_debits = _month_income_expenses(db, my)
        _, true_expenses = _month_income_expenses(db, my, exclude_investments=True)
        investment_outflow, mf_investment, other_investment = _month_investment_split(db, my)
        all_month_rows.append({
            "month": my,
            "income": income,
            "all_debits": all_debits,
            "true_expenses": true_expenses,
            "investment_outflow": investment_outflow,
            "mf_investment": mf_investment,
            "other_investment": other_investment,
            "wealth_creation": income - true_expenses,
            "raw_cash_change": income - all_debits,
        })

    complete_month_rows = [row for row in all_month_rows if row["month"] != current_month]
    month_rows = (complete_month_rows[-6:] if complete_month_rows else all_month_rows[-6:])
    past_months = [row["month"] for row in month_rows]

    months_with_data = sum(1 for row in month_rows if row["income"] > 0 or row["all_debits"] > 0)
    total_wealth_creation = sum(row["wealth_creation"] for row in month_rows)
    total_investment_outflow = sum(row["investment_outflow"] for row in month_rows)
    total_true_expenses = sum(row["true_expenses"] for row in month_rows)
    total_raw_cash_change = sum(row["raw_cash_change"] for row in month_rows)
    monthly_savings = max(total_wealth_creation / months_with_data, 0) if months_with_data else 0.0
    observed_monthly_investment_outflow = round(total_investment_outflow / months_with_data, 0) if months_with_data else 0.0
    monthly_mutual_fund_investment = round(_median_positive([row["mf_investment"] for row in month_rows]), 0)
    monthly_other_investment_outflow = round(_median_positive([row["other_investment"] for row in month_rows]), 0)
    monthly_investment_outflow = round(monthly_mutual_fund_investment + monthly_other_investment_outflow, 0)
    monthly_true_expenses = round(total_true_expenses / months_with_data, 0) if months_with_data else 0.0
    monthly_raw_cash_change = round(total_raw_cash_change / months_with_data, 0) if months_with_data else 0.0
    monthly_investment_gap = round(max(monthly_investment_outflow - monthly_savings, 0), 0)
    monthly_rate = 0.08 / 12           # 8% annualised → monthly

    data_points = [{
        "year": today.year,
        "label": str(today.year),
        "value": round(current_nw, 0),
        "base_net_worth": round(current_nw, 0),
        "base_cash": round(current_cash, 0),
        "step_net_worth": round(current_nw, 0),
        "step_cash": round(current_cash, 0),
        "base_monthly_investment": round(monthly_investment_outflow, 0),
        "step_monthly_investment": round(monthly_investment_outflow, 0),
        "base_monthly_mf_investment": round(monthly_mutual_fund_investment, 0),
        "step_monthly_mf_investment": round(monthly_mutual_fund_investment, 0),
        "monthly_other_investment_outflow": round(monthly_other_investment_outflow, 0),
        "base_unfunded_investment": 0,
        "step_unfunded_investment": 0,
        "base_cash_shortfall": 0,
        "step_cash_shortfall": 0,
    }]

    base_invested = current_invested_assets
    base_cash = current_cash
    step_invested = current_invested_assets
    step_cash = current_cash
    base_unfunded_investment = 0.0
    step_unfunded_investment = 0.0
    base_cash_shortfall = 0.0
    step_cash_shortfall = 0.0
    base_cash_runs_out_month = None
    step_cash_runs_out_month = None

    def apply_month(
        invested: float,
        cash: float,
        desired_investment: float,
        unfunded_investment: float,
        cash_shortfall: float,
        cash_runs_out_month: int | None,
        month_number: int,
    ) -> tuple[float, float, float, float, int | None, float]:
        cash_after_living = cash + monthly_savings
        if cash_after_living < 0:
            cash_shortfall += abs(cash_after_living)
            cash_after_living = 0.0

        actual_investment = min(desired_investment, cash_after_living)
        unfunded_investment += max(desired_investment - actual_investment, 0)
        cash = max(cash_after_living - actual_investment, 0)
        if cash == 0 and cash_runs_out_month is None and (cash_after_living > 0 or desired_investment > 0):
            cash_runs_out_month = month_number

        invested = invested * (1 + monthly_rate) + actual_investment
        return invested, cash, unfunded_investment, cash_shortfall, cash_runs_out_month, actual_investment

    for month in range(1, 61):         # 60 months = 5 years
        years_elapsed = (month - 1) // 12
        step_multiplier = (1 + step_up_pct / 100) ** years_elapsed
        step_monthly_mf_investment = monthly_mutual_fund_investment * step_multiplier
        step_monthly_investment = monthly_other_investment_outflow + step_monthly_mf_investment

        (
            base_invested,
            base_cash,
            base_unfunded_investment,
            base_cash_shortfall,
            base_cash_runs_out_month,
            base_actual_investment,
        ) = apply_month(
            base_invested,
            base_cash,
            monthly_investment_outflow,
            base_unfunded_investment,
            base_cash_shortfall,
            base_cash_runs_out_month,
            month,
        )

        (
            step_invested,
            step_cash,
            step_unfunded_investment,
            step_cash_shortfall,
            step_cash_runs_out_month,
            step_actual_investment,
        ) = apply_month(
            step_invested,
            step_cash,
            step_monthly_investment,
            step_unfunded_investment,
            step_cash_shortfall,
            step_cash_runs_out_month,
            month,
        )

        if month % 12 == 0:
            yr = today.year + month // 12
            base_net_worth = base_invested + base_cash
            step_net_worth = step_invested + step_cash
            data_points.append({
                "year": yr,
                "label": str(yr),
                "value": round(base_net_worth, 0),
                "base_net_worth": round(base_net_worth, 0),
                "base_cash": round(base_cash, 0),
                "step_net_worth": round(step_net_worth, 0),
                "step_cash": round(step_cash, 0),
                "base_monthly_investment": round(monthly_investment_outflow, 0),
                "step_monthly_investment": round(step_monthly_investment, 0),
                "base_monthly_mf_investment": round(monthly_mutual_fund_investment, 0),
                "step_monthly_mf_investment": round(step_monthly_mf_investment, 0),
                "monthly_other_investment_outflow": round(monthly_other_investment_outflow, 0),
                "base_actual_monthly_investment": round(base_actual_investment, 0),
                "step_actual_monthly_investment": round(step_actual_investment, 0),
                "base_unfunded_investment": round(base_unfunded_investment, 0),
                "step_unfunded_investment": round(step_unfunded_investment, 0),
                "base_cash_shortfall": round(base_cash_shortfall, 0),
                "step_cash_shortfall": round(step_cash_shortfall, 0),
            })

    return {
        "current_net_worth":      round(current_nw, 0),
        "current_cash":           round(current_cash, 0),
        "current_invested_assets": round(current_invested_assets, 0),
        "monthly_savings_assumed": round(monthly_savings, 0),
        "monthly_true_expenses":   monthly_true_expenses,
        "monthly_investment_outflow": monthly_investment_outflow,
        "monthly_mutual_fund_investment": monthly_mutual_fund_investment,
        "monthly_other_investment_outflow": monthly_other_investment_outflow,
        "observed_monthly_investment_outflow_avg": observed_monthly_investment_outflow,
        "monthly_investment_gap":  monthly_investment_gap,
        "monthly_raw_cash_change": monthly_raw_cash_change,
        "savings_basis":           "income_minus_true_spend",
        "investment_step_up_pct":  round(step_up_pct, 2),
        "step_up_applies_to":      "mutual_funds_only",
        "base_unfunded_investment": round(base_unfunded_investment, 0),
        "step_unfunded_investment": round(step_unfunded_investment, 0),
        "base_cash_shortfall":     round(base_cash_shortfall, 0),
        "step_cash_shortfall":     round(step_cash_shortfall, 0),
        "base_cash_runs_out_month": base_cash_runs_out_month,
        "step_cash_runs_out_month": step_cash_runs_out_month,
        "annual_return_pct":      8.0,
        "projection_years":       5,
        "data_points":            data_points,
        "months_of_cashflow_data": months_with_data,
        "has_cashflow_data":       months_with_data > 0,
        "confidence":              "high" if months_with_data >= 6 else "medium" if months_with_data >= 3 else "low",
        "notes": [
            "Forecast uses current net worth from saved assets and holdings.",
            "Savings assumption is income minus true consumption spend.",
            "Investment transfers are excluded from expenses and counted as wealth creation.",
            "Cash is never allowed to go negative; unaffordable planned investments are shown as a funding gap.",
            "Import at least 3-6 months of statements for a meaningful savings trend.",
        ] if months_with_data == 0 else [
            f"Monthly savings uses {months_with_data} imported month(s) of cashflow history: {', '.join(past_months)}.",
            "It excludes transactions categorized as Investments & Savings from expenses.",
            "Planned investing uses the median recent completed month, so one-off stock or wealth-platform transfers do not become a monthly commitment.",
            f"Step-up path increases only detected mutual fund/SIP outflow by {step_up_pct:g}% each year; other investment transfers remain flat.",
            "Cash is floored at zero; any excess investment plan is reported as unfunded instead of negative cash.",
            "Projection assumes an 8% annual return compounded monthly.",
        ],
    }


@router.post("/snapshot")
def take_wealth_snapshot(db: Session = Depends(get_db)):
    """
    Capture a net-worth snapshot for the current month and persist it.
    Call this at month-end (or manually) to build historical_wealth data.
    Upserts — safe to call multiple times in the same month.
    """
    today      = date.today()
    month_year = f"{today.year}-{today.month:02d}"

    manual   = {a.asset_type: a.value for a in db.query(ManualAsset).all()}
    mf_val   = sum(h.value for h in db.query(MFHolding).all())
    stk_val  = sum(
        (_safe(s.current_price) or _safe(s.avg_price)) * s.quantity
        for s in db.query(StockHolding).all()
    )
    gold_g   = manual.get("GOLD_GRAMS", 0)
    gold_val = _gold_value_inr(gold_g) if gold_g > 0 else 0

    liquid   = manual.get("BANK", 0)
    invested = round(mf_val + stk_val + manual.get("EPF", 0) +
                     manual.get("PPF", 0) + manual.get("NPS", 0) +
                     manual.get("FD", 0) + manual.get("REAL_ESTATE", 0) + gold_val, 2)
    total    = round(liquid + invested, 2)

    snap = db.query(HistoricalWealth).filter_by(month_year=month_year).first()
    if snap:
        snap.total_net_worth = total
        snap.total_liquid    = liquid
        snap.total_invested  = invested
        snap.updated_at      = datetime.utcnow()
    else:
        db.add(HistoricalWealth(
            month_year      = month_year,
            total_net_worth = total,
            total_liquid    = liquid,
            total_invested  = invested,
        ))

    for asset_type, value in _asset_values_simple(db).items():
        existing = db.query(AssetSnapshot).filter_by(month_year=month_year, asset_type=asset_type).first()
        if existing:
            existing.value = value
            existing.updated_at = datetime.utcnow()
        else:
            db.add(AssetSnapshot(month_year=month_year, asset_type=asset_type, value=value, source="month_close"))
    db.commit()

    return {
        "status":     "ok",
        "month_year": month_year,
        "total_net_worth": total,
        "total_liquid":    liquid,
        "total_invested":  invested,
    }


@router.post("/import-statement")
async def import_statement(
    file:    UploadFile = File(...),
    account: str        = Form(default="Bank Account"),
    no_llm:  bool       = Form(default=False),
    db: Session          = Depends(get_db),
):
    """
    Upload a bank statement CSV and import transactions into the DB.
    Returns { imported, skipped, errors }.
    """
    # Add scripts/ dir to path so we can import import_csv
    scripts_dir = str(Path(__file__).resolve().parents[1] / "scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)

    try:
        from import_bank_statement import import_csv
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"Import module unavailable: {e}")

    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are supported")

    # Save upload to a temp file
    contents = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        try:
            result = await import_csv(
                csv_path       = tmp_path,
                account_source = account,
                skip_duplicates= True,
                use_llm        = not no_llm,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not import statement: {e}")
    finally:
        os.unlink(tmp_path)

    if result.get("parsed_rows", 0) == 0:
        raise HTTPException(
            status_code=400,
            detail=(
                "No valid transactions found. Please upload a CSV with date, description, "
                "and debit/credit or amount/type columns."
            ),
        )

    latest_balance = result.get("latest_statement_balance")
    latest_balance_date = result.get("latest_statement_balance_date")
    bank_balance_updated = False
    needs_commit = False

    if latest_balance is not None:
        latest_balance_value = _safe(latest_balance)
        if latest_balance_value > 0:
            note_date = latest_balance_date or "latest statement row"
            note = f"Auto-updated from {account} statement ending {note_date}"
            bank_asset = db.query(ManualAsset).filter_by(asset_type="BANK").first()
            if bank_asset:
                bank_asset.value = latest_balance_value
                bank_asset.notes = note
                bank_asset.updated_at = datetime.utcnow()
            else:
                db.add(ManualAsset(
                    asset_type="BANK",
                    value=latest_balance_value,
                    notes=note,
                ))
            bank_balance_updated = True
            needs_commit = True

    month_year = date.today().strftime("%Y-%m")
    current_month_tx_count = db.query(Transaction).filter(
        func.strftime("%Y-%m", Transaction.date) == month_year
    ).count()
    if current_month_tx_count > 0:
        close = db.query(MonthClose).filter_by(month_year=month_year).first()
        if not close:
            close = MonthClose(month_year=month_year)
            db.add(close)
        close.bank_statement_imported = True
        close.updated_at = datetime.utcnow()
        needs_commit = True

    if needs_commit:
        db.commit()

    return {
        "status":   "ok",
        "imported": result["imported"],
        "skipped":  result["skipped"],
        "errors":   result["errors"],
        "parsed_rows": result.get("parsed_rows", 0),
        "earliest_date": result.get("earliest_date"),
        "latest_date": result.get("latest_date"),
        "date_span_days": result.get("date_span_days", 0),
        "unique_months": result.get("unique_months", []),
        "latest_statement_balance": latest_balance,
        "latest_statement_balance_date": latest_balance_date,
        "bank_balance_updated": bank_balance_updated,
        "current_month_transactions": current_month_tx_count,
        "current_month_marked_complete": current_month_tx_count > 0,
        "filename": file.filename,
    }
