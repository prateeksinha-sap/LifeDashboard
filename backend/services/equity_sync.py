from __future__ import annotations

import io
import os
import re
import asyncio
from datetime import date, datetime
from typing import Any
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database.models import EquitySecurityClassification, FundPortfolioStock, MFHolding


AMFI_CATEGORY_PAGE = "https://www.amfiindia.com/otherdata/categorisation-of-stocks"
AMFI_FALLBACK_URL = "https://www.amfiindia.com/Themes/Theme1/downloads/AverageMarketCapitalization31Dec2025.xlsx"
MFDATA_BASE_URL = "https://mfdata.in"
USER_AGENT = "LifeDashboard/1.0 (+local personal finance dashboard)"

EQUITY_CATEGORIES = ("Large Cap", "Mid Cap", "Small Cap")


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text or text in {"-", "nan", "None"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _normalise_category(value: Any) -> str | None:
    text = str(value or "").lower().replace("-", " ").replace("_", " ")
    if "large" in text and "mid" not in text:
        return "Large Cap"
    if "mid" in text:
        return "Mid Cap"
    if "small" in text or "micro" in text:
        return "Small Cap"
    return None


def _normalise_symbol(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    if not text or text in {"-", "NAN", "NONE"}:
        return None
    text = re.sub(r"\.(NS|BO)$", "", text)
    if "-" in text:
        base, suffix = text.split("-", 1)
        if suffix in {"T", "XT", "BE", "BZ", "EQ", "SM", "ST"} and base:
            text = base
    return text


def _normalise_isin(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    return text if text and text not in {"-", "NAN", "NONE"} else None


def _row_value(row: dict[str, Any], *candidates: str) -> Any:
    by_header = {_norm_header(key): value for key, value in row.items()}
    for candidate in candidates:
        key = _norm_header(candidate)
        if key in by_header:
            return by_header[key]
    return None


def _rows_from_workbook(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        header_idx = None
        for idx, raw in enumerate(raw_rows[:30]):
            headers = {_norm_header(cell) for cell in raw}
            if "companyname" in headers and ("isin" in headers or "nsesymbol" in headers):
                header_idx = idx
                break
        if header_idx is None:
            continue
        headers = [str(cell or "").strip() for cell in raw_rows[header_idx]]
        for raw in raw_rows[header_idx + 1:]:
            if not any(str(cell or "").strip() for cell in raw):
                continue
            values = list(raw) + [None] * max(0, len(headers) - len(raw))
            rows.append({headers[i]: values[i] for i in range(len(headers)) if headers[i]})
    return rows


async def _discover_amfi_url(client: httpx.AsyncClient) -> str:
    env_url = os.getenv("AMFI_MARKET_CAP_URL", "").strip()
    if env_url:
        return env_url
    try:
        response = await client.get(AMFI_CATEGORY_PAGE, headers={"User-Agent": USER_AGENT}, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        candidates: list[tuple[date, str]] = []
        for link in soup.find_all("a", href=True):
            href = str(link["href"])
            label = f"{href} {link.get_text(' ', strip=True)}".lower()
            if ".xlsx" not in href.lower():
                continue
            if "averagemarketcapitalization" not in label and "market capitalization" not in label:
                continue
            parsed_date = _date_from_amfi_label(label)
            if parsed_date:
                candidates.append((parsed_date, urljoin(AMFI_CATEGORY_PAGE, href)))
        if candidates:
            return sorted(candidates, key=lambda item: item[0], reverse=True)[0][1]
    except Exception:
        pass
    return AMFI_FALLBACK_URL


def _date_from_amfi_label(label: str) -> date | None:
    month_map = {
        "jan": 1, "january": 1,
        "jun": 6, "june": 6,
        "dec": 12, "december": 12,
    }
    match = re.search(r"(30|31)\s*(jun|june|dec|december|jan|january)\s*(20\d{2})", label, re.I)
    if not match:
        return None
    return date(int(match.group(3)), month_map[match.group(2).lower()], int(match.group(1)))


async def refresh_amfi_security_master(db: Session) -> dict[str, Any]:
    """Download the latest AMFI large/mid/small stock classification and upsert it."""
    async with httpx.AsyncClient(follow_redirects=True) as client:
        url = await _discover_amfi_url(client)
        response = await client.get(url, headers={"User-Agent": USER_AGENT}, timeout=60)
        response.raise_for_status()

    rows = _rows_from_workbook(response.content)
    if not rows:
        raise ValueError("AMFI file did not contain a usable stock classification table.")

    db.query(EquitySecurityClassification).filter(
        EquitySecurityClassification.source.like("AMFI:%")
    ).delete(synchronize_session=False)
    db.flush()

    imported = updated = skipped = 0
    for idx, row in enumerate(rows, start=1):
        name = str(_row_value(row, "Company name", "Stock name", "Security name", "Name") or "").strip()
        isin = _normalise_isin(_row_value(row, "ISIN", "ISIN No", "ISIN Number"))
        symbol = _normalise_symbol(_row_value(row, "NSE Symbol", "Symbol", "Ticker", "BSE Symbol"))
        sector = str(_row_value(row, "Sector", "Industry", "Basic Industry") or "").strip() or None
        category = _normalise_category(_row_value(row, "Category", "Classification", "Market Cap Category"))
        if not category:
            rank = _safe_float(_row_value(row, "Sr. No.", "Rank", "Serial No", "S No")) or idx
            if rank <= 100:
                category = "Large Cap"
            elif rank <= 250:
                category = "Mid Cap"
            else:
                category = "Small Cap"

        if not name or category not in EQUITY_CATEGORIES:
            skipped += 1
            continue

        existing = None
        if symbol:
            existing = db.query(EquitySecurityClassification).filter_by(symbol=symbol).first()
        if not existing and isin:
            existing = db.query(EquitySecurityClassification).filter_by(isin=isin).first()
        if not existing:
            existing = (
                db.query(EquitySecurityClassification)
                .filter_by(security_name=name, symbol=symbol, isin=isin)
                .first()
            )

        if existing:
            existing.security_name = name
            existing.symbol = symbol or existing.symbol
            existing.isin = isin or existing.isin
            existing.category = category
            existing.sector = sector or existing.sector
            existing.source = f"AMFI:{url}"
            existing.updated_at = datetime.utcnow()
            updated += 1
        else:
            db.add(EquitySecurityClassification(
                security_name=name,
                symbol=symbol,
                isin=isin,
                category=category,
                sector=sector,
                source=f"AMFI:{url}",
            ))
            imported += 1

    db.commit()
    return {
        "status": "ok",
        "source": url,
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "total": db.query(EquitySecurityClassification).count(),
    }


def _token_set(value: str) -> set[str]:
    stop = {
        "fund", "direct", "regular", "plan", "growth", "option", "idcw",
        "the", "and", "of", "scheme", "formerly", "known", "as",
    }
    return {
        token
        for token in re.findall(r"[a-z0-9]+", value.lower())
        if token and token not in stop
    }


def _score_candidate(holding: MFHolding, candidate: dict[str, Any]) -> float:
    score = 0.0
    holding_isin = _normalise_isin(holding.isin)
    candidate_isin = _normalise_isin(candidate.get("isin"))
    if holding_isin and candidate_isin == holding_isin:
        score += 100
    holding_tokens = _token_set(holding.scheme_name)
    candidate_tokens = _token_set(str(candidate.get("name") or ""))
    if holding_tokens and candidate_tokens:
        score += 40 * (len(holding_tokens & candidate_tokens) / len(holding_tokens | candidate_tokens))
    if holding.amc and str(candidate.get("amc_name") or "").lower().startswith(holding.amc.split()[0].lower()):
        score += 8
    scheme_lower = holding.scheme_name.lower()
    plan_type = str(candidate.get("plan_type") or "").lower()
    if plan_type and plan_type in scheme_lower:
        score += 5
    return score


async def _mfdata_get(client: httpx.AsyncClient, path: str, **params: Any) -> dict[str, Any]:
    response = await client.get(
        f"{MFDATA_BASE_URL}{path}",
        params={k: v for k, v in params.items() if v not in (None, "")},
        headers={"User-Agent": USER_AGENT},
        timeout=45,
    )
    if response.status_code == 429:
        retry_after = _safe_float(response.headers.get("retry-after")) or 65
        await asyncio.sleep(min(retry_after + 2, 90))
        response = await client.get(
            f"{MFDATA_BASE_URL}{path}",
            params={k: v for k, v in params.items() if v not in (None, "")},
            headers={"User-Agent": USER_AGENT},
            timeout=45,
        )
    response.raise_for_status()
    payload = response.json()
    if payload.get("status") != "success":
        raise ValueError(str(payload))
    return payload


async def _resolve_mfdata_family(client: httpx.AsyncClient, holding: MFHolding) -> dict[str, Any] | None:
    candidates: list[dict[str, Any]] = []

    if holding.isin:
        try:
            payload = await _mfdata_get(client, "/api/v1/search", q=holding.isin)
            candidates.extend(payload.get("data") or [])
        except Exception:
            pass

    if not candidates and holding.scheme_code:
        try:
            payload = await _mfdata_get(
                client,
                f"/api/v1/schemes/{holding.scheme_code}",
                fields="name,family_id,category,amc_name,isin,plan_type",
            )
            data = payload.get("data") or {}
            if data:
                data["amfi_code"] = str(holding.scheme_code)
                candidates.append(data)
        except Exception:
            pass

    if not candidates:
        try:
            payload = await _mfdata_get(client, "/api/v1/search", q=holding.scheme_name)
            candidates.extend(payload.get("data") or [])
        except Exception:
            pass

    if not candidates:
        return None

    ranked = sorted(candidates, key=lambda item: _score_candidate(holding, item), reverse=True)
    best = ranked[0]
    if _score_candidate(holding, best) < 35:
        return None
    return best if best.get("family_id") else None


def _fund_may_have_stock_exposure(scheme_name: str, category: str | None = None) -> bool:
    text = f"{scheme_name or ''} {category or ''}".lower()
    negative = (
        "liquid", "overnight", "ultra short", "short duration", "low duration",
        "money market", "gilt", "debt", "bond", "floating", "floater", "gold savings",
        "gold fund", "silver", "arbitrage",
    )
    positive = (
        "equity", "flexi", "large", "mid", "small", "contra", "elss", "nifty",
        "index", "multicap", "multi cap", "business cycle", "sector", "thematic",
        "digital", "resources", "energy", "mnc", "services", "innovation",
        "fund of funds", "fof",
    )
    if any(word in text for word in positive):
        return True
    return not any(word in text for word in negative)


def _parse_month_start(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m").date()
    except ValueError:
        return None


async def sync_mfdata_fund_portfolios(db: Session, limit: int | None = None) -> dict[str, Any]:
    """Resolve current MF holdings to mfdata families and import latest equity portfolios."""
    holdings = db.query(MFHolding).order_by(MFHolding.value.desc()).all()
    unique: dict[str, MFHolding] = {}
    for holding in holdings:
        key = holding.isin or holding.scheme_name
        unique.setdefault(key, holding)

    selected = list(unique.values())
    if limit:
        selected = selected[: max(1, limit)]

    resolved = imported_rows = updated_funds = no_equity = missing = errors = 0
    details: list[dict[str, Any]] = []
    async with httpx.AsyncClient(follow_redirects=True) as client:
        for holding in selected:
            detail = {
                "scheme_name": holding.scheme_name,
                "value": round(float(holding.value or 0), 2),
                "status": "pending",
            }
            try:
                if not _fund_may_have_stock_exposure(holding.scheme_name):
                    no_equity += 1
                    detail["status"] = "no_stock_exposure_expected"
                    details.append(detail)
                    continue
                match = await _resolve_mfdata_family(client, holding)
                if not match:
                    missing += 1
                    detail["status"] = "not_resolved"
                    details.append(detail)
                    continue
                resolved += 1
                family_id = int(match["family_id"])
                if not _fund_may_have_stock_exposure(holding.scheme_name, str(match.get("category") or "")):
                    no_equity += 1
                    detail.update({
                        "status": "no_stock_exposure_expected",
                        "family_id": family_id,
                        "matched_name": match.get("name"),
                    })
                    details.append(detail)
                    continue
                try:
                    payload = await _mfdata_get(
                        client,
                        f"/api/v1/families/{family_id}/holdings",
                        holding_type="equity",
                        max_results=1000,
                    )
                except httpx.HTTPStatusError as exc:
                    if exc.response.status_code == 404:
                        no_equity += 1
                        detail.update({
                            "status": "no_portfolio_data",
                            "family_id": family_id,
                            "matched_name": match.get("name"),
                        })
                        details.append(detail)
                        continue
                    raise
                data = payload.get("data") or {}
                equity_holdings = data.get("equity_holdings") or []
                if not equity_holdings:
                    no_equity += 1
                    detail.update({
                        "status": "no_equity_holdings",
                        "family_id": family_id,
                        "matched_name": match.get("name"),
                    })
                    details.append(detail)
                    continue

                db.query(FundPortfolioStock).filter_by(scheme_name=holding.scheme_name).delete()
                as_of = _parse_month_start(data.get("month"))
                for item in equity_holdings:
                    weight = _safe_float(item.get("weight_pct"))
                    stock_name = str(item.get("stock_name") or item.get("name") or "").strip()
                    if not stock_name or weight <= 0:
                        continue
                    db.add(FundPortfolioStock(
                        scheme_name=holding.scheme_name,
                        amc=holding.amc or match.get("amc_name"),
                        stock_name=stock_name,
                        symbol=_normalise_symbol(item.get("symbol")),
                        isin=_normalise_isin(item.get("isin")),
                        category=None,
                        sector=str(item.get("sector") or "").strip() or None,
                        weight_pct=weight,
                        as_of_date=as_of,
                        source=f"mfdata.in family={family_id} month={data.get('month') or 'latest'}",
                    ))
                    imported_rows += 1
                updated_funds += 1
                detail.update({
                    "status": "synced",
                    "family_id": family_id,
                    "matched_name": match.get("name"),
                    "month": data.get("month"),
                    "equity_holdings": len(equity_holdings),
                    "equity_pct": data.get("equity_pct"),
                })
                details.append(detail)
                db.commit()
            except Exception as exc:
                db.rollback()
                errors += 1
                detail["status"] = "error"
                detail["error"] = str(exc)[:240]
                details.append(detail)

    return {
        "status": "ok",
        "provider": "mfdata.in",
        "funds_seen": len(selected),
        "funds_resolved": resolved,
        "funds_synced": updated_funds,
        "funds_without_equity": no_equity,
        "funds_unresolved": missing,
        "portfolio_rows_imported": imported_rows,
        "errors": errors,
        "details": details,
    }


async def sync_equity_lookthrough(db: Session, limit: int | None = None) -> dict[str, Any]:
    security_master = await refresh_amfi_security_master(db)
    fund_portfolios = await sync_mfdata_fund_portfolios(db, limit=limit)
    return {
        "status": "ok",
        "security_master": security_master,
        "fund_portfolios": fund_portfolios,
    }
